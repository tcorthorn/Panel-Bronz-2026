from bronz_app import views
import os
from django.contrib.admin.views.decorators import staff_member_required # type: ignore
from django.shortcuts import render, redirect
from django.urls import reverse
from django.contrib import messages
from django.http import JsonResponse
from bronz_app.utils import regenerar_resumenes_credito_debito
import pandas as pd
import openpyxl

from datetime import date, datetime

PANEL_YEAR_CHOICES = (2025, 2026)


def _get_default_panel_year() -> int:
    today = date.today()
    if today.year in PANEL_YEAR_CHOICES:
        return today.year
    return PANEL_YEAR_CHOICES[0]


def get_panel_year(request) -> int:
    year = request.session.get('panel_year')
    if year is None or int(year) not in PANEL_YEAR_CHOICES:
        year = _get_default_panel_year()
        request.session['panel_year'] = year
    return int(year)


def get_panel_date_range(request):
    year = get_panel_year(request)
    start_date = date(year, 1, 1)
    today = date.today()
    if year == today.year:
        end_date = today
    else:
        end_date = date(year, 12, 31)
    return year, start_date, end_date


def regenerate_financial_tables(start_date, end_date):
    regenerar_ventas_consulta(start_date=start_date, end_date=end_date)  # CORREGIDO
    poblar_movimientos_unificados_debito(start_date=start_date, end_date=end_date)
    poblar_movimientos_unificados_credito(start_date=start_date, end_date=end_date)
    regenerar_resumenes_credito_debito()


from openpyxl.utils.dataframe import dataframe_to_rows
from django.http import HttpResponse
from .eerr import generar_estado_resultados
from .utils_balance import obtener_matriz_balance
from .balance_utils import obtener_matriz_dict_balance
from .utils import regenerar_ventas_consulta
from django.views.decorators.csrf import csrf_exempt
from bronz_app.models import ResumenCredito, ResumenDebito, MovimientoUnificadoCredito, MovimientoUnificadoDebito
from bronz_app.scripts.export_a_excel import main as export_a_excel_script
from django.views.decorators.http import require_POST
from django.shortcuts import redirect, render
from django.contrib import messages
from django.views.decorators.http import require_POST
from bronz_app.scripts.import_ajuste_inventario import main as import_ajuste_inventario_script
from bronz_app.scripts.import_asientos_contables import main as import_asientos_contables_script
from bronz_app.scripts.import_catalogo import main as import_catalogo_script
from bronz_app.scripts.import_balance_inicial import main as import_balance_inicial_script
from bronz_app.scripts.import_entrada_productos import main as import_entrada_productos_script
from bronz_app.scripts.import_inventario_inicial import main as import_inventario_inicial_script
from bronz_app.scripts.import_sueldos import main as import_sueldos_script
from bronz_app.scripts.import_ventas import main as import_ventas_script
from bronz_app.scripts.import_envios import main as import_envios_script
from bronz_app.scripts.import_otros_gastos import main as import_otros_gastos_script
from bronz_app.scripts.import_shopify_orders import main as import_shopify_orders_script
from django.contrib.auth.decorators import login_required
from datetime import date, datetime
from decimal import Decimal
from collections import defaultdict
from django.db.models.functions import TruncMonth
from django.db.models import Sum
from .models import ResumenMensual
from .utils import (
    regenerar_ventas_consulta,
    poblar_movimientos_unificados_debito,
    poblar_movimientos_unificados_credito,
    regenerar_resumenes_credito_debito,
)
from django.views.decorators.csrf import csrf_exempt




@require_POST
def set_panel_year(request):
    year_value = request.POST.get('panel_year')
    try:
        year_int = int(year_value)
    except (TypeError, ValueError):
        year_int = _get_default_panel_year()
    if year_int not in PANEL_YEAR_CHOICES:
        year_int = _get_default_panel_year()
    request.session['panel_year'] = year_int
    next_url = request.POST.get('next') or request.META.get('HTTP_REFERER') or reverse('home')
    return redirect(next_url)


def home(request):
    year, start_date, end_date = get_panel_date_range(request)
    context = {
        'panel_year': year,
        'panel_years': PANEL_YEAR_CHOICES,
        'panel_start_date': start_date,
        'panel_end_date': end_date,
    }
    return render(request, "bronz_app/home.html", context)


def listado_union_credito(request):
    total = MovimientoUnificadoCredito.objects.count()

    # <-- Aquí va la lógica para devolver JSON si se pide:
    if request.GET.get('ver') == 'json':
        data = list(
            MovimientoUnificadoCredito.objects.values(
                'fecha', 'cta_credito', 'monto_credito', 'texto_coment', 'tabla_origen'
            )
        )
        return JsonResponse(data, safe=False)

    # Si no llega ?ver=json, sigue mostrando la plantilla con el botón
    return render(request, 'bronz_app/union_credito.html', {
        'total_registros': total
    })

def cargar_union_credito(request):
    """
    Esta vista se invoca cuando el usuario hace clic en “Cargar movimientos unificados”.
    -- Llama a poblar_movimientos_unificados() para insertar/actualizar.
    -- Luego redirige a listado_union y muestra un mensaje “Cargado con éxito”.
    """
    try:
        poblar_movimientos_unificados_credito()
        messages.success(request, f"Tabla 'movimiento_unificado' actualizada: ahora hay {MovimientoUnificadoCredito.objects.count()} registros.")
    except Exception as e:
        # Captura cualquier error y muestra mensaje de error
        messages.error(request, f"Error al poblar movimientos unificados: {e}")

    return redirect(reverse('listado_union_credito'))

def listado_union_debito(request):
    """
    Similar a listado_union, pero para la tabla de débitos.
    """
    total = MovimientoUnificadoDebito.objects.count()
    return render(request, 'bronz_app/union_debito.html', {
        'total_registros': total
    })

def cargar_union_debito(request):
    """
    Dispara la carga para la tabla movemento_unificado_debito.
    """
    try:
        poblar_movimientos_unificados_debito()
        messages.success(request, f"Tabla 'movimiento_unificado_debito' actualizada: ahora hay {MovimientoUnificadoDebito.objects.count()} registros.")
    except Exception as e:
        messages.error(request, f"Error al poblar movimientos débito: {e}")

    return redirect(reverse('listado_union_debito'))

# ——————————————————————————————————————————————————————————————
#  VENTAS CONSULTA
# ————————————————————————————————————————————————————————

@staff_member_required
def regenerar_consulta_view(request):
    if request.method == "POST":
        regenerar_ventas_consulta()
        return render(request, "bronz_app/confirmacion.html", {"mensaje": "✅ VentasConsulta regenerada correctamente."})

    return render(request, "bronz_app/confirmacion.html", {"mensaje": None})   

# ——————————————————————————————————————————————————————————————
#  PAGINA INICIO
# ————————————————————————————————————————————————————————

@staff_member_required
def pagina_inicio(request):
    return render(request, "bronz_app/inicio.html")

@staff_member_required
def regenerar_resumenes_view(request):
    mensaje = None
    if request.method == "POST":
        total_creditos, total_debitos = regenerar_resumenes_credito_debito()
        mensaje = f"✅ Regenerado: {total_creditos} créditos y {total_debitos} débitos."

    return render(request, "bronz_app/resumen_regenerar.html", {"mensaje": mensaje})

# ——————————————————————————————————————————————————————————————
#  EXPORTAR A EXCEL SUMA CREDITOS SUMA DEBITOS A GESTION BRONZ PARA BALANCE
# ————————————————————————————————————————————————————————

@staff_member_required
@csrf_exempt
def exportar_resumen_excel(request):
    if request.method == "POST":
        ruta_libro = r"C:\Users\tcort\OneDrive\BRONZ\Django\Otros\GestionBronzX.xlsm"
        hoja_destino = "Datos"
        if not os.path.exists(ruta_libro):
            messages.error(request, f"❌ Archivo no encontrado: {ruta_libro}")
            return redirect('home')

        # --- Resúmenes y movimientos ---
        df_credito = pd.DataFrame(list(ResumenCredito.objects.values("cuenta_credito", "total_credito")))
        df_debito = pd.DataFrame(list(ResumenDebito.objects.values("cuenta_debito", "total_debito")))
        df_unif_credito = pd.DataFrame(list(MovimientoUnificadoCredito.objects.values("fecha", "cta_credito", "monto_credito")))
        df_unif_debito  = pd.DataFrame(list(MovimientoUnificadoDebito.objects.values("fecha", "cta_debito", "monto_debito")))

        filas_credito = len(df_credito)
        filas_debito = len(df_debito)
        filas_unif_credito = len(df_unif_credito)
        filas_unif_debito  = len(df_unif_debito)

        # --- Abrir Excel ---
        libro = openpyxl.load_workbook(ruta_libro, keep_vba=True)
        if hoja_destino not in libro.sheetnames:
            messages.error(request, f"❌ Hoja '{hoja_destino}' no encontrada en el libro.")
            return redirect('home')
        hoja = libro[hoja_destino]

        # Limpiar áreas destino
        for fila in hoja.iter_rows(min_row=2, min_col=1, max_col=3):   # A-C
            for celda in fila: celda.value = None
        for fila in hoja.iter_rows(min_row=2, min_col=5, max_col=7):   # E-G
            for celda in fila: celda.value = None
        for fila in hoja.iter_rows(min_row=2, min_col=10, max_col=13): # J-M
            for celda in fila: celda.value = None

        # Escribir movimientos
        for i, row in df_unif_debito.iterrows():
            hoja[f"A{i+2}"] = row["fecha"]
            hoja[f"B{i+2}"] = row["cta_debito"]
            hoja[f"C{i+2}"] = row["monto_debito"]
        for i, row in df_unif_credito.iterrows():
            hoja[f"E{i+2}"] = row["fecha"]
            hoja[f"F{i+2}"] = row["cta_credito"]
            hoja[f"G{i+2}"] = row["monto_credito"]

        # Escribir resumenes
        max_filas = max(filas_credito, filas_debito)
        for i in range(max_filas):
            hoja[f"J{i+2}"] = df_debito.iloc[i]["cuenta_debito"] if i < filas_debito else ""
            hoja[f"K{i+2}"] = df_debito.iloc[i]["total_debito"] if i < filas_debito else ""
            hoja[f"L{i+2}"] = df_credito.iloc[i]["cuenta_credito"] if i < filas_credito else ""
            hoja[f"M{i+2}"] = df_credito.iloc[i]["total_credito"] if i < filas_credito else ""

        libro.save(ruta_libro)

        # Un solo mensaje, sólo HTML seguro
        messages.success(
            request,
            (
                "✅ Exportación completada.<br>"
                f"📊 Créditos exportados (resumen): {filas_credito} <br>"
                f"📊 Débitos exportados (resumen): {filas_debito} <br>"
                f"📤 Movimientos crédito: {filas_unif_credito} <br>"
                f"📤 Movimientos débito: {filas_unif_debito}"
            )
        )
        return redirect('home')

    # Redirige siempre a home si la petición no es POST
    return redirect('home')

# ——————————————————————————————————————————————————————————————
#  EXPORTAR A EXCEL TODOS CREDITOS Y TODOS DEBITOS A GESTION BRONZ PARA BALANCE
# ————————————————————————————————————————————————————————

@require_POST
def export_a_excel_view(request):
    msg = exportar_resumen_excel()
    if "Error" in msg or "error" in msg or "🧨" in msg:
        messages.error(request, msg)
    else:
        messages.success(request, msg)
    return redirect('home')

 # ——————————————————————————————————————————————————————————————
# VISTAS EN HOME IMPORTACIONES
# ————————————————————————————————————————————————————————   




@require_POST
def import_ajuste_inventario(request):
    try:
        mensaje = import_ajuste_inventario_script()
        messages.success(request, mensaje)
    except Exception as e:
        messages.error(request, f"Error al importar Ajuste de Inventario: {str(e)}")
    return redirect('home')

@require_POST
def import_asientos_contables(request):
    try:
        msg = import_asientos_contables_script()
        messages.success(request, msg)
    except Exception as e:
        messages.error(request, f"Error al importar Asientos Contables: {str(e)}")
    return redirect('home')

@require_POST
def import_catalogo(request):
    try:
        resultado = import_catalogo_script()
        if resultado and "no hay" in resultado.lower():
            messages.info(request, resultado)
        else:
            messages.success(request, resultado)
    except Exception as e:
        messages.error(request, f"Error al importar Catálogo: {str(e)}")
    return redirect('home')

@require_POST
def import_balance_inicial(request):
    try:
        msg = import_balance_inicial_script()
        messages.success(request, msg)
    except Exception as e:
        messages.error(request, f"Error al importar Balance Inicial: {str(e)}")
    return redirect('home')

@require_POST
def import_entrada_productos(request):
    try:
        msg = import_entrada_productos_script()
        messages.success(request, msg)
    except Exception as e:
        messages.error(request, f"Error al importar Entradas de Productos: {str(e)}")
    return redirect('home')

@require_POST
def import_inventario_inicial(request):
    try:
        msg = import_inventario_inicial_script()
        messages.success(request, msg)
    except Exception as e:
        messages.error(request, f"Error al importar Inventario Inicial: {str(e)}")
    return redirect('home')

@require_POST
def import_sueldos(request):
    try:
        msg = import_sueldos_script()
        messages.success(request, msg)
    except Exception as e:
        messages.error(request, f"Error al importar Sueldos: {str(e)}")
    return redirect('home')

@require_POST
def import_ventas(request):
    try:
        msg = import_ventas_script()
        messages.success(request, msg)
    except Exception as e:
        messages.error(request, f"Error al importar Ventas: {str(e)}")
    return redirect('home')

@require_POST
def import_envios(request):
    try:
        mensaje = import_envios_script()
        messages.success(request, mensaje)
    except Exception as e:
        messages.error(request, f"Error al importar Envíos: {str(e)}")
    return redirect('home')

@require_POST
def import_otros_gastos(request):
    try:
        msg = import_otros_gastos_script()
        messages.success(request, msg)
    except Exception as e:
        messages.error(request, f"Error al importar Otros Gastos: {str(e)}")
    return redirect('home')

@require_POST
def import_shopify_orders(request):
    """Importa órdenes desde un archivo CSV de Shopify."""
    try:
        msg = import_shopify_orders_script()
        messages.success(request, msg)
    except Exception as e:
        messages.error(request, f"Error al importar Órdenes Shopify: {str(e)}")
    return redirect('home')

# ——————————————————————————————————————————————————————————————
# DASHBOARD SHOPIFY
# ——————————————————————————————————————————————————————————————
from bronz_app.shopify_dashboard import get_shopify_dashboard_data

def shopify_dashboard(request):
    """Dashboard de análisis de ventas Shopify."""
    data = get_shopify_dashboard_data()
    return render(request, 'bronz_app/shopify_dashboard.html', {'data': data})

@login_required
def procesar_todo(request):
    if request.method == 'POST':
        try:
            # Ejecutar en orden las 4 funciones utilitarias
            poblar_movimientos_unificados_credito()
            poblar_movimientos_unificados_debito()
            regenerar_resumenes_credito_debito()
            regenerar_ventas_consulta()
            messages.success(request, "¡Todos los procesos ejecutados correctamente en orden:! (VentasConsulta, Unión Créditos, Unión Débitos y Resúmenes)")
        except Exception as e:
            messages.error(request, f"Error al ejecutar procesamiento total: {e}")
    return redirect('home')


# ——————————————————————————————————————————————————————————————
# VISTAS EN HOME PROCESAR DATOS
# ————————————————————————————————————————————————————————   

from django.contrib.auth.decorators import login_required

@login_required
def procesar_ventas_consulta(request):
    if request.method == 'POST':
        regenerar_ventas_consulta()
        messages.success(request, "¡Tabla VentasConsulta regenerada!")
    return redirect('home')

@login_required
def procesar_union_credito(request):
    if request.method == 'POST':
        poblar_movimientos_unificados_credito()
        messages.success(request, "¡Unión de Créditos procesada con éxito!")
    return redirect('home')

@login_required
def procesar_union_debito(request):
    if request.method == 'POST':
        poblar_movimientos_unificados_debito()
        messages.success(request, "¡Unión de Débitos procesada con éxito!")
    return redirect('home')

@login_required
def procesar_resumenes(request):
    if request.method == 'POST':
        regenerar_resumenes_credito_debito()
        messages.success(request, "¡Resúmenes de crédito y débito regenerados!")
    return redirect('home')

# ——————————————————————————————————————————————————————————————
# VISTAS EN HOME INVENTARIO
# ————————————————————————————————————————————————————————   

from django.contrib import messages
from django.core.management import call_command
import io
import re

@login_required
def procesar_inventario(request):
    if request.method == 'POST':
        out = io.StringIO()
        try:
            call_command('inventario', stdout=out)
            resultado = out.getvalue()
            # Limpiar ANSI:
            ansi_escape = re.compile(r'\x1B[@-_][0-?]*[ -/]*[@-~]')
            resultado_limpio = ansi_escape.sub('', resultado)
            messages.success(request, f"<pre>{resultado_limpio}</pre>")
        except Exception as e:
            messages.error(request, f"Error al procesar inventario: {e}")
    return redirect('home')

from django.http import HttpResponse
from openpyxl import Workbook
from .models import Catalogo, InventarioInicial, EntradaProductos, Envios, Ventas, AjusteInventario
from django.db import models
from django.urls import reverse

def exportar_inventario_actual(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventario Actual"
    ws.append(['SKU', 'Categoría', 'Producto', 'Stock', 'Bodega', 'Ingresos', 'Envios', 'Ventas', 'Ajustes'])
    for obj in Catalogo.objects.all():
        ini = InventarioInicial.objects.filter(sku=obj.sku).first()
        stock = ini.stock if ini else 0
        bodega = ini.bodega if ini else 0
        ingresos = EntradaProductos.objects.filter(sku__sku=obj.sku).aggregate(total=models.Sum('cantidad_ingresada'))['total'] or 0
        envios = Envios.objects.filter(sku__sku=obj.sku).aggregate(total=models.Sum('cantidad'))['total'] or 0
        ventas = Ventas.objects.filter(sku__sku=obj.sku).aggregate(total=models.Sum('cantidad'))['total'] or 0
        ajustes = AjusteInventario.objects.filter(sku__sku=obj.sku).aggregate(total=models.Sum('cantidad'))['total'] or 0

        ws.append([
            obj.sku, obj.categoria, obj.producto, stock, bodega,
            ingresos, envios, ventas, ajustes
        ])
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=InventarioActual.xlsx'
    wb.save(response)
    return response

# ——————————————————————————————————————————————————————————————
# TABLAS PARA INVENTARIO
# ——————————————————————————————————————————————————————————————


from django.shortcuts import render
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.urls import reverse
from .models import Catalogo, InventarioInicial, EntradaProductos, Envios, Ventas, AjusteInventario
from django.db import models

def inventario_actual(request):
    q = request.GET.get('q', '').strip()
    sort = request.GET.get('sort', 'sku')
    direction = request.GET.get('dir', 'asc')

    columnas_ordenables = {
        'sku': 'sku',
        'categoria': 'categoria',
        'producto': 'producto',
        'inicial': None,
        'bodega': None,
        'ingresos': None,
        'envios': None,
        'ventas': None,
        'ajustes': None,
        'en_oficina': None,
        'en_bodega': None,
        'ajuste_ventas': None,
        'total': None,
    }

    todos = Catalogo.objects.all()
    if q:
        todos = todos.filter(
            models.Q(sku__icontains=q) |
            models.Q(categoria__icontains=q) |
            models.Q(producto__icontains=q)
        )

    if columnas_ordenables.get(sort):
        orden = columnas_ordenables[sort]
        if direction == 'desc':
            orden = '-' + orden
        todos = todos.order_by(orden)

        # ---- Prepara ventas por SKU para ajuste ----
    ventas_por_sku = {}
    for obj in todos:
        ventas = Ventas.objects.filter(sku__sku=obj.sku).aggregate(total=models.Sum('cantidad'))['total'] or 0
        ventas_por_sku[obj.sku] = ventas

    productos = []
    for obj in todos:
        ini = InventarioInicial.objects.filter(sku=obj.sku).first()
        inicial = ini.stock if ini else 0
        bodega = ini.bodega if ini else 0
        ingresos = EntradaProductos.objects.filter(sku__sku=obj.sku).aggregate(total=models.Sum('cantidad_ingresada'))['total'] or 0
        envios = Envios.objects.filter(sku__sku=obj.sku).aggregate(total=models.Sum('cantidad'))['total'] or 0
        ventas = ventas_por_sku.get(obj.sku, 0)
        ajustes = AjusteInventario.objects.filter(sku__sku=obj.sku).aggregate(total=models.Sum('cantidad'))['total'] or 0

        # Lógica de ajuste de ventas
        ajuste_ventas = ventas
        if obj.sku == 'BB0001' or obj.sku == 'BB0002':
            ajuste_ventas += ventas_por_sku.get('BB0003', 0) + ventas_por_sku.get('BB0012', 0)
        elif obj.sku == 'BB0003':
            ajuste_ventas -= ventas_por_sku.get('BB0003', 0)
        elif obj.sku == 'BB0009' or obj.sku == 'BB0010':
            ajuste_ventas += ventas_por_sku.get('BB0012', 0)
        # El resto solo sus propias ventas

        en_oficina = inicial + ingresos - envios
        # NUEVA LÓGICA EN BODEGA:
        en_bodega = bodega + envios - ajuste_ventas
        total = en_oficina + en_bodega - ajustes

        productos.append({
            'sku': obj.sku,
            'categoria': obj.categoria,
            'producto': obj.producto,
            'inicial': inicial,
            'bodega': bodega,
            'ingresos': ingresos,
            'envios': envios,
            'ventas': ventas,
            'ajustes': ajustes,
            'en_oficina': en_oficina,
            'en_bodega': en_bodega,
            'ajuste_ventas': ajuste_ventas,
            'total': total,
        })


    # Ordena si corresponde
    if columnas_ordenables.get(sort) is None and productos and sort in productos[0]:
        reverse_sort = (direction == 'desc')
        productos = sorted(productos, key=lambda x: x[sort], reverse=reverse_sort)

    # Totales
    global_total_inicial = sum(p['inicial'] for p in productos)
    global_total_bodega = sum(p['bodega'] for p in productos)
    global_total_ingresos = sum(p['ingresos'] for p in productos)
    global_total_envios = sum(p['envios'] for p in productos)
    global_total_ventas = sum(p['ventas'] for p in productos)
    global_total_ajustes = sum(p['ajustes'] for p in productos)
    global_total_en_oficina = sum(p['en_oficina'] for p in productos)
    global_total_en_bodega = sum(p['en_bodega'] for p in productos)
    global_total_ajuste_ventas = sum(p['ajuste_ventas'] for p in productos)
    global_total_total = sum(p['total'] for p in productos)

    paginator = Paginator(productos, 10)
    page = request.GET.get('page')
    try:
        productos_page = paginator.page(page)
    except PageNotAnInteger:
        productos_page = paginator.page(1)
    except EmptyPage:
        productos_page = paginator.page(paginator.num_pages)

    total_inicial = sum(p['inicial'] for p in productos_page.object_list)
    total_bodega = sum(p['bodega'] for p in productos_page.object_list)
    total_ingresos = sum(p['ingresos'] for p in productos_page.object_list)
    total_envios = sum(p['envios'] for p in productos_page.object_list)
    total_ventas = sum(p['ventas'] for p in productos_page.object_list)
    total_ajustes = sum(p['ajustes'] for p in productos_page.object_list)
    total_en_oficina = sum(p['en_oficina'] for p in productos_page.object_list)
    total_en_bodega = sum(p['en_bodega'] for p in productos_page.object_list)
    total_ajuste_ventas = sum(p['ajuste_ventas'] for p in productos_page.object_list)
    total_total = sum(p['total'] for p in productos_page.object_list)

    url_excel = reverse('exportar_inventario_actual')
    return render(request, 'bronz_app/inventario.html', {
        'productos_page': productos_page,
        'q': q,
        'url_excel': url_excel,
        'sort': sort,
        'direction': direction,
        'total_inicial': total_inicial,
        'total_bodega': total_bodega,
        'total_ingresos': total_ingresos,
        'total_envios': total_envios,
        'total_ventas': total_ventas,
        'total_ajustes': total_ajustes,
        'total_en_oficina': total_en_oficina,
        'total_en_bodega': total_en_bodega,
        'total_ajuste_ventas': total_ajuste_ventas,
        'total_total': total_total,
        'global_total_inicial': global_total_inicial,
        'global_total_bodega': global_total_bodega,
        'global_total_ingresos': global_total_ingresos,
        'global_total_envios': global_total_envios,
        'global_total_ventas': global_total_ventas,
        'global_total_ajustes': global_total_ajustes,
        'global_total_en_oficina': global_total_en_oficina,
        'global_total_en_bodega': global_total_en_bodega,
        'global_total_ajuste_ventas': global_total_ajuste_ventas,
        'global_total_total': global_total_total,
    })



# ——————————————————————————————————————————————————————————————
# BALANCE
# ——————————————————————————————————————————————————————————————

from datetime import date  # <-- Importa SOLO la clase date
from .utils import (
    regenerar_ventas_consulta,
    poblar_movimientos_unificados_debito,
    poblar_movimientos_unificados_credito,
    regenerar_resumenes_credito_debito,
)

from django.shortcuts import render
from django.http import HttpResponse
from .models import ResumenDebito, ResumenCredito
import pandas as pd
from .cod_cuentas_balance import balance_rows

def intdot(val):
    """Formatea números para mostrar miles con punto."""
    try:
        val_float = float(val)
        val_int = int(round(val_float))
        return f"{val_int:,}".replace(",", ".")
    except Exception:
        return ""

def balance_view(request):

    # Ejecutar siempre los procesos previos antes de mostrar el balance
    year, start_date, end_date = get_panel_date_range(request)
    regenerate_financial_tables(start_date, end_date)

    debitos_dict = {d.cuenta_debito: float(d.total_debito) for d in ResumenDebito.objects.all()}
    creditos_dict = {c.cuenta_credito: float(c.total_credito) for c in ResumenCredito.objects.all()}

    matriz_balance = []
    fecha_corte = date.today().strftime("%Y-%m-%d")
    total_debito = total_credito = total_saldo_deudor = total_saldo_acreedor = 0
    total_activo = total_pasivo = total_perdidas = total_ganancias = 0

    for fila in balance_rows:
        codigo = fila['codigo']
        nombre = fila['nombre']
        debito = debitos_dict.get(codigo, 0)
        credito = creditos_dict.get(codigo, 0)
        saldo_deudor = saldo_acreedor = activo = pasivo = perdidas = ganancias = 0

        if 1010100 <= codigo <= 2040000:
            saldo_deudor = debito - credito if debito > credito else 0
            saldo_acreedor = credito - debito if credito > debito else 0
            activo = saldo_deudor
            pasivo = saldo_acreedor
        elif 3010100 <= codigo <= 3030300:
            saldo_deudor = debito - credito if debito > credito else 0
            saldo_acreedor = credito - debito if credito > debito else 0
            perdidas = saldo_deudor
            ganancias = saldo_acreedor

        matriz_balance.append({
            'codigo': codigo,
            'nombre': nombre,
            'debito': intdot(debito),
            'credito': intdot(credito),
            'saldo_deudor': intdot(saldo_deudor),
            'saldo_acreedor': intdot(saldo_acreedor),
            'activo': intdot(activo),
            'pasivo': intdot(pasivo),
            'perdidas': intdot(perdidas),
            'ganancias': intdot(ganancias)
        })

        # Acumula totales
        total_debito += debito
        total_credito += credito
        total_saldo_deudor += saldo_deudor
        total_saldo_acreedor += saldo_acreedor
        total_activo += activo
        total_pasivo += pasivo
        total_perdidas += perdidas
        total_ganancias += ganancias

    # -- Agrega fila: Utilidad (pérdida) del Ejercicio
    def resta_positiva(a, b):
        return max(0, a - b)

    utilidad_activo = resta_positiva(total_pasivo, total_activo)
    utilidad_pasivo = resta_positiva(total_activo, total_pasivo)
    utilidad_perdidas = resta_positiva(total_ganancias, total_perdidas)
    utilidad_ganancias = resta_positiva(total_perdidas, total_ganancias)

    utilidad = {
        'debito': '',
        'credito': '',
        'saldo_deudor': '',
        'saldo_acreedor': '',
        'activo': intdot(utilidad_activo),
        'pasivo': intdot(utilidad_pasivo),
        'perdidas': intdot(utilidad_perdidas),
        'ganancias': intdot(utilidad_ganancias)
    }

    # Flags para pintar en rojo
    utilidad_pasivo_rojo = total_activo < total_pasivo
    utilidad_perdidas_rojo = total_perdidas > total_ganancias

    # -- Agrega fila: SUMAS TOTALES (Totales + Utilidad)
    sumas_totales = {
        'debito': intdot(total_debito),
        'credito': intdot(total_credito),
        'saldo_deudor': intdot(total_saldo_deudor),
        'saldo_acreedor': intdot(total_saldo_acreedor),
        'activo': intdot(total_activo + utilidad_activo),
        'pasivo': intdot(total_pasivo + utilidad_pasivo),
        'perdidas': intdot(total_perdidas + utilidad_perdidas),
        'ganancias': intdot(total_ganancias + utilidad_ganancias)
    }

    totales = {
        'debito': intdot(total_debito),
        'credito': intdot(total_credito),
        'saldo_deudor': intdot(total_saldo_deudor),
        'saldo_acreedor': intdot(total_saldo_acreedor),
        'activo': intdot(total_activo),
        'pasivo': intdot(total_pasivo),
        'perdidas': intdot(total_perdidas),
        'ganancias': intdot(total_ganancias)
    }

    # Exportar a Excel si es solicitado
    if request.GET.get("export") == "excel":
        df = pd.DataFrame([
            {**f, **{k: v.replace('.', '') for k, v in f.items() if k not in ('codigo', 'nombre')}}
            for f in matriz_balance
        ])
        response = HttpResponse(content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename="balance.xlsx"'
        df.to_excel(response, index=False)
        return response

    return render(request, "bronz_app/balance.html", {
        'matriz_balance': matriz_balance,
        'totales': totales,
        'utilidad': utilidad,
        'sumas_totales': sumas_totales,
        'utilidad_pasivo_rojo': utilidad_pasivo_rojo,
        'utilidad_perdidas_rojo': utilidad_perdidas_rojo,
        'fecha_corte': fecha_corte,  # <--- Aquí pasas la fecha de hoy
        'panel_year': year,
        'panel_start_date': start_date,
        'panel_end_date': end_date,
    })

# ——————————————————————————————————————————————————————————————
# RESUMEN BALANCE
# ——————————————————————————————————————————————————————————————

import json
from django.shortcuts import render
# Asegúrate de importar tus modelos de balance, igual que haces para balance_view

def resumen_balance_view(request):

    # Ejecutar siempre los procesos previos antes de mostrar el balance
    year, start_date, end_date = get_panel_date_range(request)
    regenerate_financial_tables(start_date, end_date)

    # 1. Prepara los datos "matriz_dict"
    # Debe ser: {'A:1010100': 123, 'P:1010100': 0, ...}
    # Puedes obtenerlo de tu lógica existente para balance_view
    # Aquí te doy un ejemplo muy simple:
    matriz_dict = {}
    # Supón que tienes lista de filas balance_rows y los valores de cada columna
    # Puedes reutilizar tu código de balance_view:
    from .models import ResumenDebito, ResumenCredito
    from .cod_cuentas_balance import balance_rows

    debitos_dict = {d.cuenta_debito: float(d.total_debito) for d in ResumenDebito.objects.all()}
    creditos_dict = {c.cuenta_credito: float(c.total_credito) for c in ResumenCredito.objects.all()}
    for fila in balance_rows:
        codigo = str(fila['codigo'])
        debito = debitos_dict.get(int(codigo), 0)
        credito = creditos_dict.get(int(codigo), 0)
        # Calcula las columnas igual que tu balance_view
        saldo_deudor = debito - credito if debito > credito else 0
        saldo_acreedor = credito - debito if credito > debito else 0
        # Según rango, asigna a A, P, Pe, G
        if 1010100 <= int(codigo) <= 2040000:
            matriz_dict[f'A:{codigo}'] = saldo_deudor
            matriz_dict[f'P:{codigo}'] = saldo_acreedor
        elif 3010100 <= int(codigo) <= 3010300:
            matriz_dict[f'Pe:{codigo}'] = saldo_deudor
            matriz_dict[f'G:{codigo}'] = saldo_acreedor

    # 2. Renderiza el template y pasa matriz_js
    return render(request, "bronz_app/resumen_balance.html", {
        'matriz_js': json.dumps(matriz_dict),
        'panel_year': year,
        'panel_start_date': start_date,
        'panel_end_date': end_date,
    })

# ——————————————————————————————————————————————————————————————
# BALANCE SEGUN FECHA
# ——————————————————————————————————————————————————————————————

from django.shortcuts import render
from django.http import HttpResponse
from .models import MovimientoUnificadoCredito, MovimientoUnificadoDebito
from .cod_cuentas_balance import balance_rows
from datetime import datetime
import pandas as pd

def balance_segun_fecha_view(request):

    year, start_date, default_end_date = get_panel_date_range(request)

    fecha_corte = request.GET.get('fecha_corte')
    fecha_corte_dt = None
    fecha_corte_str = ''

    if fecha_corte:
        try:
            fecha_corte_dt = datetime.strptime(fecha_corte, "%Y-%m-%d").date()   # <- cambio aquí
            fecha_corte_str = fecha_corte_dt.strftime("%Y-%m-%d")                # <- para mostrar igual en el título
        except Exception:
            fecha_corte_dt = None
            fecha_corte_str = ''

    if not fecha_corte_dt:
        return render(request, "bronz_app/balance_segun_fecha.html", {
            'panel_year': year,
            'panel_start_date': start_date,
            'panel_end_date': default_end_date,
        })

    range_start = date(fecha_corte_dt.year, 1, 1)
    effective_end = min(fecha_corte_dt, default_end_date)
    if range_start > effective_end:
        range_start = date(effective_end.year, 1, 1)
    regenerate_financial_tables(range_start, effective_end)

    debitos = MovimientoUnificadoDebito.objects.filter(fecha__gte=range_start, fecha__lte=effective_end)
    creditos = MovimientoUnificadoCredito.objects.filter(fecha__gte=range_start, fecha__lte=effective_end)

    debitos_dict = {}
    for d in debitos:
        debitos_dict[d.cta_debito] = debitos_dict.get(d.cta_debito, 0) + float(d.monto_debito)
    creditos_dict = {}
    for c in creditos:
        creditos_dict[c.cta_credito] = creditos_dict.get(c.cta_credito, 0) + float(c.monto_credito)

    matriz_balance = []
    total_debito = total_credito = total_saldo_deudor = total_saldo_acreedor = 0
    total_activo = total_pasivo = total_perdidas = total_ganancias = 0

    def intdot(val):
        try:
            val_float = float(val)
            val_int = int(round(val_float))
            return f"{val_int:,}".replace(",", ".")
        except Exception:
            return ""

    for fila in balance_rows:
        codigo = fila['codigo']
        nombre = fila['nombre']
        debito = debitos_dict.get(codigo, 0)
        credito = creditos_dict.get(codigo, 0)
        saldo_deudor = saldo_acreedor = activo = pasivo = perdidas = ganancias = 0

        if 1010100 <= codigo <= 2040000:
            saldo_deudor = debito - credito if debito > credito else 0
            saldo_acreedor = credito - debito if credito > debito else 0
            activo = saldo_deudor
            pasivo = saldo_acreedor
        elif 3010100 <= codigo <= 3010300:
            saldo_deudor = debito - credito if debito > credito else 0
            saldo_acreedor = credito - debito if credito > debito else 0
            perdidas = saldo_deudor
            ganancias = saldo_acreedor

        matriz_balance.append({
            'codigo': codigo,
            'nombre': nombre,
            'debito': intdot(debito),
            'credito': intdot(credito),
            'saldo_deudor': intdot(saldo_deudor),
            'saldo_acreedor': intdot(saldo_acreedor),
            'activo': intdot(activo),
            'pasivo': intdot(pasivo),
            'perdidas': intdot(perdidas),
            'ganancias': intdot(ganancias)
        })

        total_debito += debito
        total_credito += credito
        total_saldo_deudor += saldo_deudor
        total_saldo_acreedor += saldo_acreedor
        total_activo += activo
        total_pasivo += pasivo
        total_perdidas += perdidas
        total_ganancias += ganancias

    def resta_positiva(a, b):
        return max(0, a - b)

    utilidad_activo = resta_positiva(total_pasivo, total_activo)
    utilidad_pasivo = resta_positiva(total_activo, total_pasivo)
    utilidad_perdidas = resta_positiva(total_ganancias, total_perdidas)
    utilidad_ganancias = resta_positiva(total_perdidas, total_ganancias)

    utilidad = {
        'debito': '',
        'credito': '',
        'saldo_deudor': '',
        'saldo_acreedor': '',
        'activo': intdot(utilidad_activo),
        'pasivo': intdot(utilidad_pasivo),
        'perdidas': intdot(utilidad_perdidas),
        'ganancias': intdot(utilidad_ganancias)
    }

    utilidad_pasivo_rojo = total_activo < total_pasivo
    utilidad_perdidas_rojo = total_perdidas > total_ganancias

    sumas_totales = {
        'debito': intdot(total_debito),
        'credito': intdot(total_credito),
        'saldo_deudor': intdot(total_saldo_deudor),
        'saldo_acreedor': intdot(total_saldo_acreedor),
        'activo': intdot(total_activo + utilidad_activo),
        'pasivo': intdot(total_pasivo + utilidad_pasivo),
        'perdidas': intdot(total_perdidas + utilidad_perdidas),
        'ganancias': intdot(total_ganancias + utilidad_ganancias)
    }

    totales = {
        'debito': intdot(total_debito),
        'credito': intdot(total_credito),
        'saldo_deudor': intdot(total_saldo_deudor),
        'saldo_acreedor': intdot(total_saldo_acreedor),
        'activo': intdot(total_activo),
        'pasivo': intdot(total_pasivo),
        'perdidas': intdot(total_perdidas),
        'ganancias': intdot(total_ganancias)
    }

    # Exportar a Excel si es solicitado
    if request.GET.get("export") == "excel":
        df = pd.DataFrame([
            {**f, **{k: v.replace('.', '') for k, v in f.items() if k not in ('codigo', 'nombre')}}
            for f in matriz_balance
        ])
        response = HttpResponse(content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = f'attachment; filename="balance_segun_fecha_{fecha_corte_str}.xlsx"'
        df.to_excel(response, index=False)
        return response

    return render(request, "bronz_app/balance.html", {
        'matriz_balance': matriz_balance,
        'totales': totales,
        'utilidad': utilidad,
        'sumas_totales': sumas_totales,
        'utilidad_pasivo_rojo': utilidad_pasivo_rojo,
        'utilidad_perdidas_rojo': utilidad_perdidas_rojo,
        'fecha_corte': fecha_corte_str,
        'panel_year': year,
        'panel_start_date': range_start,
        'panel_end_date': effective_end,
    })


# ————————————————————————————————————————————————————————
# RESUMEN BALANCE SEGÚN FECHA
# ————————————————————————————————————————————————————————
from django.shortcuts import render
import json
from .models import MovimientoUnificadoCredito, MovimientoUnificadoDebito
from .cod_cuentas_balance import balance_rows

def resumen_balance_segun_fecha_view(request):

    year, start_date, default_end = get_panel_date_range(request)

    fecha_corte = request.GET.get('fecha_corte')
    matriz_dict = {}
    fecha_corte_dt = None
    effective_end = default_end
    range_start = start_date

    if fecha_corte:
        from datetime import datetime
        try:
            fecha_corte_dt = datetime.strptime(fecha_corte, "%Y-%m-%d").date()
        except ValueError:
            fecha_corte_dt = None
        if fecha_corte_dt:
            range_start = date(fecha_corte_dt.year, 1, 1)
            effective_end = min(fecha_corte_dt, default_end)
            if range_start > effective_end:
                range_start = date(effective_end.year, 1, 1)
            regenerate_financial_tables(range_start, effective_end)
            debitos = MovimientoUnificadoDebito.objects.filter(fecha__gte=range_start, fecha__lte=effective_end)
            creditos = MovimientoUnificadoCredito.objects.filter(fecha__gte=range_start, fecha__lte=effective_end)

            debitos_dict = {}
            for d in debitos:
                debitos_dict[d.cta_debito] = debitos_dict.get(d.cta_debito, 0) + float(d.monto_debito)
            creditos_dict = {}
            for c in creditos:
                creditos_dict[c.cta_credito] = creditos_dict.get(c.cta_credito, 0) + float(c.monto_credito)
            for fila in balance_rows:
                codigo = str(fila['codigo'])
                debito = debitos_dict.get(int(codigo), 0)
                credito = creditos_dict.get(int(codigo), 0)
                saldo_deudor = debito - credito if debito > credito else 0
                saldo_acreedor = credito - debito if credito > debito else 0
                if 1010100 <= int(codigo) <= 2040000:
                    matriz_dict[f'A:{codigo}'] = saldo_deudor
                    matriz_dict[f'P:{codigo}'] = saldo_acreedor
                elif 3010100 <= int(codigo) <= 3010300:
                    matriz_dict[f'Pe:{codigo}'] = saldo_deudor
                    matriz_dict[f'G:{codigo}'] = saldo_acreedor

    return render(request, "bronz_app/resumen_balance_segun_fecha.html", {
        'matriz_js': json.dumps(matriz_dict),
        'fecha_corte': fecha_corte,
        'panel_year': year,
        'panel_start_date': range_start,
        'panel_end_date': effective_end,
    })


# ——————————————————————————————————————————————————————————————
# RESUMEN FINANCIERO - VIEW
# ——————————————————————————————————————————————————————————————

import json
from django.shortcuts import render

def resumen_financiero(request):
    # Procesos previos (puedes incluir solo los necesarios para el financiero)
    year, start_date, end_date = get_panel_date_range(request)
    regenerate_financial_tables(start_date, end_date)

    matriz_dict = {}
    from .models import ResumenDebito, ResumenCredito
    from .cod_cuentas_balance import balance_rows   # o tu archivo resumen_financiero_rows si tienes uno diferente

    debitos_dict = {d.cuenta_debito: float(d.total_debito) for d in ResumenDebito.objects.all()}
    creditos_dict = {c.cuenta_credito: float(c.total_credito) for c in ResumenCredito.objects.all()}

    for fila in balance_rows:   # Si tienes otro set de rows para el financiero, cámbialo aquí
        codigo = str(fila['codigo'])
        debito = debitos_dict.get(int(codigo), 0)
        credito = creditos_dict.get(int(codigo), 0)
        saldo_deudor = debito - credito if debito > credito else 0
        saldo_acreedor = credito - debito if credito > debito else 0
        if 1010100 <= int(codigo) <= 2040000:
            matriz_dict[f'A:{codigo}'] = saldo_deudor
            matriz_dict[f'P:{codigo}'] = saldo_acreedor
        elif 3010100 <= int(codigo) <= 3010300:
            matriz_dict[f'Pe:{codigo}'] = saldo_deudor
            matriz_dict[f'G:{codigo}'] = saldo_acreedor

    return render(request, "bronz_app/resumen_financiero.html", {
        'matriz_js': json.dumps(matriz_dict),
        'panel_year': year,
        'panel_start_date': start_date,
        'panel_end_date': end_date,
    })


# ——————————————————————————————————————————————————————————————
# RESUMEN FINANCIERO SEGUN FECHA CORTE- VIEW
# ——————————————————————————————————————————————————————————————

from django.shortcuts import render
import json
from .models import MovimientoUnificadoCredito, MovimientoUnificadoDebito
from .cod_cuentas_balance import balance_rows
from datetime import datetime

def resumen_financiero_segun_fecha_view(request):

    year, start_date, default_end = get_panel_date_range(request)

    fecha_corte = request.GET.get('fecha_corte')
    matriz_dict = {}
    fecha_corte_dt = None
    effective_end = default_end
    range_start = start_date

    if fecha_corte:
        try:
            # Permite ambos formatos de fecha: dd-mm-yyyy y yyyy-mm-dd
            try:
                fecha_corte_dt = datetime.strptime(fecha_corte, "%d-%m-%Y").date()
            except ValueError:
                fecha_corte_dt = datetime.strptime(fecha_corte, "%Y-%m-%d").date()
            effective_end = min(fecha_corte_dt, default_end)
            range_start = date(fecha_corte_dt.year, 1, 1)
            if range_start > effective_end:
                range_start = date(effective_end.year, 1, 1)
            regenerate_financial_tables(range_start, effective_end)
            debitos = MovimientoUnificadoDebito.objects.filter(fecha__gte=range_start, fecha__lte=effective_end)
            creditos = MovimientoUnificadoCredito.objects.filter(fecha__gte=range_start, fecha__lte=effective_end)
            debitos_dict = {}
            for d in debitos:
                debitos_dict[d.cta_debito] = debitos_dict.get(d.cta_debito, 0) + float(d.monto_debito)
            creditos_dict = {}
            for c in creditos:
                creditos_dict[c.cta_credito] = creditos_dict.get(c.cta_credito, 0) + float(c.monto_credito)
            for fila in balance_rows:
                codigo = str(fila['codigo'])
                debito = debitos_dict.get(int(codigo), 0)
                credito = creditos_dict.get(int(codigo), 0)
                saldo_deudor = debito - credito if debito > credito else 0
                saldo_acreedor = credito - debito if credito > debito else 0
                if 1010100 <= int(codigo) <= 2040000:
                    matriz_dict[f'A:{codigo}'] = saldo_deudor
                    matriz_dict[f'P:{codigo}'] = saldo_acreedor
                elif 3010100 <= int(codigo) <= 3010300:
                    matriz_dict[f'Pe:{codigo}'] = saldo_deudor
                    matriz_dict[f'G:{codigo}'] = saldo_acreedor
        except Exception:
            pass

    return render(request, "bronz_app/resumen_financiero_segun_fecha.html", {
        'matriz_js': json.dumps(matriz_dict),
        'fecha_corte': fecha_corte,
        'panel_year': year,
        'panel_start_date': range_start,
        'panel_end_date': effective_end,
    })

# views.py

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from bronz_app.resumen_financiero import RESUMEN_ACTIVO, RESUMEN_PASIVO, RESUMEN_RESULTADO
from bronz_app.utils_financiero import eval_formula

def exportar_excel_resumen_financiero(request):
    # Supón que ya generaste matriz_dict en tu view principal según la fecha de corte
    matriz_dict = ...  # aquí debes obtenerla (¡esto es parte de tu lógica!)

    # 1. Calcula resultados en arrays:
    resultados_activo = {}
    for row in RESUMEN_ACTIVO:
        val = eval_formula(row["formula"], matriz_dict, resultados_activo)
        resultados_activo[str(row["linea"])] = val
        row["resultado"] = val

    resultados_resultado = {}
    for row in RESUMEN_RESULTADO:
        val = eval_formula(row["formula"], matriz_dict, resultados_resultado)
        resultados_resultado[str(row["linea"])] = val
        row["resultado"] = val

    resultados_pasivo = {}
    for row in RESUMEN_PASIVO:
        val = eval_formula(
            row["formula"], matriz_dict, resultados_pasivo,
            resultado_lookup=resultados_resultado,
            activo_lookup=resultados_activo
        )
        resultados_pasivo[str(row["linea"])] = val
        row["resultado"] = val

    # 2. Arma la tabla con tres bloques en la misma hoja:
    resumen = [
        ["Línea", "Cuenta Activo", "Monto Activo",
         "Línea", "Cuenta Pasivo", "Monto Pasivo",
         "Línea", "Cuenta Resultado", "Monto Resultado"]
    ]
    max_rows = max(len(RESUMEN_ACTIVO), len(RESUMEN_PASIVO), len(RESUMEN_RESULTADO))
    for i in range(max_rows):
        fila_activo = RESUMEN_ACTIVO[i] if i < len(RESUMEN_ACTIVO) else {"linea":"", "cuenta":"", "resultado":""}
        fila_pasivo = RESUMEN_PASIVO[i] if i < len(RESUMEN_PASIVO) else {"linea":"", "cuenta":"", "resultado":""}
        fila_resultado = RESUMEN_RESULTADO[i] if i < len(RESUMEN_RESULTADO) else {"linea":"", "cuenta":"", "resultado":""}
        resumen.append([
            fila_activo["linea"], fila_activo["cuenta"], fila_activo["resultado"],
            fila_pasivo["linea"], fila_pasivo["cuenta"], fila_pasivo["resultado"],
            fila_resultado["linea"], fila_resultado["cuenta"], fila_resultado["resultado"]
        ])

    # 3. Exporta a Excel PRO
    wb = Workbook()
    ws = wb.active
    ws.title = "Resumen Financiero"

    # Cabecera pro
    header_font = Font(bold=True, size=13, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F81BD")
    alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin', color='B7B7B7'),
        right=Side(style='thin', color='B7B7B7'),
        top=Side(style='thin', color='B7B7B7'),
        bottom=Side(style='thin', color='B7B7B7')
    )
    ws.append(resumen[0])
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = alignment
        cell.border = thin_border
    ws.freeze_panes = 'A2'

    fill_even = PatternFill("solid", fgColor="E9F1FB")
    fill_odd = PatternFill("solid", fgColor="FFFFFF")
    for row_idx, row in enumerate(resumen[1:], start=2):
        ws.append(row)
        for col_idx, cell_value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal="right")
            else:
                cell.alignment = Alignment(horizontal="left")
            cell.border = thin_border
            cell.fill = fill_even if row_idx % 2 == 0 else fill_odd

    for col_idx in range(1, len(resumen[0]) + 1):
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in ws[get_column_letter(col_idx)])
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 2, 45)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="ResumenFinanciero.xlsx"'
    wb.save(response)
    return response

#----------------------------------------------
# RESUMEN MENSUAL
#---------------------------------------------




# ---------------------------------------------
# RESUMEN MENSUAL Helpers
# ---------------------------------------------
def to_month_start(value):
    """Normaliza 'mes' a date(YYYY,MM,1) soportando date, datetime o str."""
    if value is None:
        return None
    if isinstance(value, date) and not isinstance(value, datetime):
        return date(value.year, value.month, 1)
    if isinstance(value, datetime):
        d = value.date()
        return date(d.year, d.month, 1)
    if isinstance(value, str):
        s = value.strip()
        for fmt in ("%Y-%m", "%Y-%m-%d"):
            try:
                dt = datetime.strptime(s, fmt)
                return date(dt.year, dt.month, 1)
            except ValueError:
                continue
        # fallback flexible: "YYYY-M[-D]"
        try:
            y, m, *_ = s.split("-")
            return date(int(y), int(m), 1)
        except Exception:
            pass
    raise ValueError(f"Formato de 'mes' no reconocido: {value!r}")

def add_month(d: date) -> date:
    return date(d.year + (d.month == 12), 1 if d.month == 12 else d.month + 1, 1)


# ---------------------------------------------
# DASHBOARD / RESUMEN MENSUAL
# ---------------------------------------------
def dashboard(request):
    """
    - Regenera orígenes -> escribe ResumenMensual.
    - Muestra desde ENERO del año vigente si hay datos del año en curso,
      si no, desde ENERO del último año con datos, hasta el último mes con datos.
    - Rellena meses faltantes con 0 y calcula utilidad_acumulada (en memoria).
    """
    year, start_date, end_date = get_panel_date_range(request)

    # 1) Procesos previos (igual que la app que sí te funciona)
    try:
        regenerate_financial_tables(start_date, end_date)
    except Exception as e:
        # No detengas el dashboard por errores de regeneración
        messages.warning(request, f"Advertencia en regeneración: {e}")

    # 2) Lee ResumenMensual ya actualizado
    qs = (
        ResumenMensual.objects
        .filter(mes__gte=start_date, mes__lte=end_date)
        .values('mes', 'ventas', 'costos', 'utilidad')
        .order_by('mes')
    )
    rows = list(qs)

    # 3) Agrega por mes normalizado
    data_by_month = {}
    for r in rows:
        try:
            mes = to_month_start(r.get('mes'))
        except ValueError:
            continue
        if mes is None:
            continue
        ventas   = Decimal(r.get('ventas') or 0)
        costos   = Decimal(r.get('costos') or 0)
        utilidad = r.get('utilidad')
        utilidad = Decimal(utilidad) if utilidad is not None else (ventas - costos)

        if mes not in data_by_month:
            data_by_month[mes] = {'mes': mes, 'ventas': ventas, 'costos': costos, 'utilidad': utilidad}
        else:
            data_by_month[mes]['ventas']   += ventas
            data_by_month[mes]['costos']   += costos
            data_by_month[mes]['utilidad'] += utilidad

    if not data_by_month:
        contexto = {
            'resumenes': [],
            'totales': {'ventas': Decimal('0'), 'costos': Decimal('0'), 'utilidad': Decimal('0')},
            'chart_data': {'labels': [], 'ventas': [], 'costos': [], 'utilidad': []},
            'panel_year': year,
            'panel_start_date': start_date,
            'panel_end_date': end_date,
        }
        return render(request, 'bronz_app/resumen_mensual.html', contexto)

    # 4) Rango de meses mostrado
    months_sorted   = sorted(data_by_month.keys())
    end             = min(months_sorted[-1], end_date) if months_sorted else end_date
    start           = start_date
    if end < start:
        end = start

    # 5) Reconstrucción continua + acumulado
    resumenes = []
    cur = start
    running = Decimal('0')
    while cur <= end:
        base = data_by_month.get(cur, {'mes': cur, 'ventas': Decimal('0'), 'costos': Decimal('0'), 'utilidad': Decimal('0')})
        reg = {
            'mes': cur,
            'ventas': Decimal(base['ventas']),
            'costos': Decimal(base['costos']),
            'utilidad': Decimal(base['utilidad']),
        }
        running += reg['utilidad']
        reg['utilidad_acumulada'] = running
        resumenes.append(reg)
        cur = add_month(cur)

    # 6) Totales + Chart.js
    totales = {
        'ventas':   sum((r['ventas']   for r in resumenes), Decimal('0')),
        'costos':   sum((r['costos']   for r in resumenes), Decimal('0')),
        'utilidad': sum((r['utilidad'] for r in resumenes), Decimal('0')),
    }
    chart_data = {
        'labels':   [f"{r['mes'].year:04d}-{r['mes'].month:02d}" for r in resumenes],
        'ventas':   [float(r['ventas'])   for r in resumenes],
        'costos':   [float(r['costos'])   for r in resumenes],
        'utilidad': [float(r['utilidad']) for r in resumenes],
    }

    # DEBUG temporal
    print("DEBUG last label:", chart_data['labels'][-1:] if chart_data['labels'] else None)

    return render(request, 'bronz_app/resumen_mensual.html', {
        'resumenes': resumenes,
        'totales': totales,
        'chart_data': chart_data,
        'panel_year': year,
        'panel_start_date': start_date,
        'panel_end_date': end_date,
    })


# ---------------------------------------------
# ACTUALIZAR RESUMEN MENSUAL (igual a la app OK)
# ---------------------------------------------
CUENTAS_VENTAS = [3010101, 3010111]
CUENTA_COSTO = [3010200, 3010201, 3010202, 3010203,3010205,3010211,3010212,3010213,3010214,3010215,3010216, 3020200,3010300,3010400,3020500,3020600,3020700,3020800,3020900,3030100]

def actualizar_resumen_mensual(request):
    # Regenera orígenes
    year, start_date, end_date = get_panel_date_range(request)
    regenerate_financial_tables(start_date, end_date)

    # Recalcula saldos por mes directamente desde los movimientos
    saldos_mensuales = defaultdict(lambda: defaultdict(lambda: {'debitos': 0, 'creditos': 0}))

    debitos = MovimientoUnificadoDebito.objects.annotate(
        mes=TruncMonth('fecha')
    ).values('mes', 'cta_debito').annotate(total=Sum('monto_debito'))

    for row in debitos:
        mes = row['mes']
        cuenta = str(row['cta_debito'])
        saldos_mensuales[mes][cuenta]['debitos'] += row['total'] or 0

    creditos = MovimientoUnificadoCredito.objects.annotate(
        mes=TruncMonth('fecha')
    ).values('mes', 'cta_credito').annotate(total=Sum('monto_credito'))

    for row in creditos:
        mes = row['mes']
        cuenta = str(row['cta_credito'])
        saldos_mensuales[mes][cuenta]['creditos'] += row['total'] or 0

    CUENTA_COSTO_STR = [str(c) for c in CUENTA_COSTO]
    CUENTAS_VENTAS_STR = [str(c) for c in CUENTAS_VENTAS]

    for mes, cuentas in saldos_mensuales.items():
        ventas = sum(
            cuentas.get(c, {}).get('creditos', 0) - cuentas.get(c, {}).get('debitos', 0)
            for c in CUENTAS_VENTAS_STR
        )
        costo_venta = sum(
            cuentas.get(c, {}).get('debitos', 0) - cuentas.get(c, {}).get('creditos', 0)
            for c in CUENTA_COSTO_STR
        )
        utilidad = ventas - costo_venta
        margen_bruto = utilidad

        ResumenMensual.objects.update_or_create(
            mes=mes,
            defaults={
                'ventas': ventas,
                'costos': costo_venta,
                'utilidad': utilidad,
                'margen_bruto': margen_bruto
            }
        )

    # Actualiza utilidad acumulada en BD
    resumenes = ResumenMensual.objects.order_by('mes')
    acumulado = 0
    for r in resumenes:
        acumulado += r.utilidad or 0
        if r.utilidad_acumulada != acumulado:
            r.utilidad_acumulada = acumulado
            r.save(update_fields=['utilidad_acumulada'])

    return redirect('dashboard')

# ————————————————————————————————————————————————————————
#EXPORTAR EXCEL RESUMEN MENSUAL
# ————————————————————————————————————————————————————————

def exportar_resumen_excel(request):
    """
    Exporta el mismo rango que se ve en 'dashboard' a Excel (.xlsx).
    Si openpyxl no está disponible, cae a CSV.
    """
    resumenes = _build_resumenes_desde_enero_ultimo_anio()

    # Intentar Excel con openpyxl
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.utils import get_column_letter
        from io import BytesIO

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Resumen Mensual'

        headers = ['Mes', 'Ventas', 'Costos', 'Utilidad', 'Utilidad Acumulada']
        ws.append(headers)

        for r in resumenes:
            ws.append([
                r['mes'].strftime('%Y-%m'),
                float(r['ventas']),
                float(r['costos']),
                float(r['utilidad']),
                float(r['utilidad_acumulada']),
            ])

        # Estilos básicos
        bold = Font(bold=True)
        fill = PatternFill('solid', fgColor='EAF3FA')
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=c)
            cell.font = bold
            cell.fill = fill
            cell.alignment = Alignment(horizontal='center')

        # Formato numérico y ancho de columnas
        for row in ws.iter_rows(min_row=2, min_col=2, max_col=5):
            for cell in row:
                cell.number_format = '#,##0'
        for col in ws.columns:
            maxlen = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val = '' if cell.value is None else str(cell.value)
                maxlen = max(maxlen, len(val))
            ws.column_dimensions[col_letter].width = min(maxlen + 2, 24)

        # Respuesta HTTP
        stream = BytesIO()
        wb.save(stream)
        stream.seek(0)
        filename = f"ResumenMensual_{date.today().isoformat()}.xlsx"
        resp = HttpResponse(
            stream.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        resp['Content-Disposition'] = f'attachment; filename="{filename}"'
        return resp

    except Exception:
        # Fallback CSV
        import csv
        resp = HttpResponse(content_type='text/csv')
        resp['Content-Disposition'] = f'attachment; filename="ResumenMensual_{date.today().isoformat()}.csv"'
        writer = csv.writer(resp)
        writer.writerow(['Mes', 'Ventas', 'Costos', 'Utilidad', 'Utilidad Acumulada'])
        for r in resumenes:
            writer.writerow([
                r['mes'].strftime('%Y-%m'),
                int(r['ventas']),
                int(r['costos']),
                int(r['utilidad']),
                int(r['utilidad_acumulada']),
            ])
        return resp

# ——————————————————————————————————————————————————————————————
# RESULTADO DETALLADO POR MES
# ——————————————————————————————————————————————————————————————

from django.shortcuts import render
from .models import ResultadoMensualDetalle
from collections import defaultdict
from datetime import date
# Importa tus funciones utilitarias de procesamiento aquí
from .utils import (
    regenerar_ventas_consulta,
    poblar_movimientos_unificados_credito,
    poblar_movimientos_unificados_debito,
    regenerar_resumenes_credito_debito, 
)

CONCEPTOS_FINALES = [
    ("ventas", "Ventas"),
    ("costo_venta", "Costo de Venta"),
    ("margen_bruto", "Margen Bruto"),
    ("gastos_comercializacion", "Gastos Comercialización"),
    ("comision_plataformas", "Comision Plataformas de Pago"),
    ("gastos_marketing", "Gastos Publicidad y Marketing"),
    ("gastos_arriendos", "Gastos arriendos Comisiones Tiendas"),
    ("gastos_envios", "Gastos de Envíos Adicionales"),
    ("gastos_administracion", "Gastos de Administración"),
    ("resultado_operacional_bruto", "Resultado Operacional Bruto"),
    ("gastos_financieros", "Gastos Financieros"),
    ("depreciacion", "Depreciación del Ejercicio"),
    ("resultado_oper_neto", "Resultado Oper. Neto"),
    ("utilidad_no_operacional", "Utilidad (pérd.) No Operacional"),
    ("ajuste_monetario", "Ajuste Monetario"),
    ("impuesto_renta", "Impuesto a la Renta"),
    ("ajustes", "Ajustes"),
    ("utilidad_neta", "Utilidad Neta del Periodo"),
]

FILAS_CALCULADAS = [
    "margen_bruto",
    "resultado_operacional_bruto",
    "resultado_oper_neto",
    "utilidad_neta",
]

def tabla_resultados_mensual(request):

     # Procesos previos (puedes incluir solo los necesarios para el financiero)
    year, start_date, end_date = get_panel_date_range(request)
    regenerate_financial_tables(start_date, end_date)

    año = year
    mes_actual = end_date.month

    detalles = ResultadoMensualDetalle.objects.filter(mes__gte=start_date, mes__lte=end_date)
    datos = defaultdict(lambda: defaultdict(int))
    for d in detalles:
        datos[d.mes.month][d.concepto] = d.valor

    matriz = {c[0]: [None]*mes_actual for c in CONCEPTOS_FINALES}
    for mes in range(1, mes_actual + 1):
        base = datos[mes]
        margen_bruto = base.get("ventas", 0) - base.get("costo_venta", 0)
        gastos_suma = sum(base.get(k, 0) for k in [
            "gastos_comercializacion", "comision_plataformas","gastos_marketing", "gastos_arriendos",
            "gastos_envios", "gastos_administracion"
        ])
        resultado_operacional_bruto = margen_bruto - gastos_suma
        resultado_oper_neto = resultado_operacional_bruto - base.get("gastos_financieros", 0) - base.get("depreciacion", 0)
        utilidad_neta = (
            resultado_oper_neto
            - base.get("utilidad_no_operacional", 0)
            - base.get("ajuste_monetario", 0)
            - base.get("impuesto_renta", 0)
            - base.get("ajustes", 0)
        )
        matriz["ventas"][mes-1] = base.get("ventas")
        matriz["costo_venta"][mes-1] = base.get("costo_venta")
        matriz["margen_bruto"][mes-1] = margen_bruto
        matriz["gastos_comercializacion"][mes-1] = base.get("gastos_comercializacion")
        matriz["comision_plataformas"][mes-1] = base.get("comision_plataformas")
        matriz["gastos_marketing"][mes-1] = base.get("gastos_marketing")
        matriz["gastos_arriendos"][mes-1] = base.get("gastos_arriendos")
        matriz["gastos_envios"][mes-1] = base.get("gastos_envios")
        matriz["gastos_administracion"][mes-1] = base.get("gastos_administracion")
        matriz["resultado_operacional_bruto"][mes-1] = resultado_operacional_bruto
        matriz["gastos_financieros"][mes-1] = base.get("gastos_financieros")
        matriz["depreciacion"][mes-1] = base.get("depreciacion")
        matriz["resultado_oper_neto"][mes-1] = resultado_oper_neto
        matriz["utilidad_no_operacional"][mes-1] = base.get("utilidad_no_operacional")
        matriz["ajuste_monetario"][mes-1] = base.get("ajuste_monetario")
        matriz["impuesto_renta"][mes-1] = base.get("impuesto_renta")
        matriz["ajustes"][mes-1] = base.get("ajustes")
        matriz["utilidad_neta"][mes-1] = utilidad_neta

    MESES_ES = [
        'Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
        'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre'
    ]
    meses = MESES_ES[:mes_actual]

    filas = [
        (c[0], c[1], matriz[c[0]], sum(v for v in matriz[c[0]] if v is not None))
        for c in CONCEPTOS_FINALES
    ]

    return render(request, "bronz_app/tabla_resultados_mensual.html", {
        "año": año,
        "meses": meses,
        "filas": filas,
        "filas_calculadas": FILAS_CALCULADAS,
        "panel_year": year,
        "panel_start_date": start_date,
        "panel_end_date": end_date,
    })



from django.shortcuts import redirect
from bronz_app.utils_balance import calcular_resultados_mensuales

def actualizar_resultados_mensuales(request):
    year, start_date, end_date = get_panel_date_range(request)
    try:
        regenerate_financial_tables(start_date, end_date)
        calcular_resultados_mensuales(año=year)
        messages.success(request, f"Resultados detallados recalculados para {year}.")
    except Exception as e:
        messages.error(request, f"Error al recalcular resultados detallados: {e}")
    return redirect('tabla_resultados_mensual')


from django.shortcuts import render

def importar_datos(request):
    return render(request, "bronz_app/importar_datos.html")

# ——————————————————————————————————————————————————————————————
# RESUMEN VENTAS POR TIENDA + EXPORTACIÓN A EXCEL
# ——————————————————————————————————————————————————————————————

from django.contrib.admin.views.decorators import staff_member_required
from django.shortcuts import render
from django.http import HttpResponse
from django.utils import timezone

import pandas as pd

from .models import VentasConsulta, ResumenDebito, ResumenCredito
from .utils import (regenerar_ventas_consulta, 
)
from .cod_cuentas_balance import balance_rows


# -----------------------------
# Configuración de Tiendas
# -----------------------------
TIENDAS = [
    {"key": "online",        "nombre": "Online",         "gasto_key": "Pe:3010211"},
    {"key": "falabella",     "nombre": "Falabella",      "gasto_key": "Pe:3010213"},
    {"key": "paris",         "nombre": "Paris",          "gasto_key": "Pe:3010216"},
    {"key": "ubereats",      "nombre": "Uber Eats",      "gasto_key": "Pe:3010214"},
    {"key": "mercadolibre",  "nombre": "Mercado Libre",  "gasto_key": "Pe:3010215"},
    {"key": "ventamanual",   "nombre": "Venta Manual",   "gasto_key": None},
    {"key": "otros",         "nombre": "Otros",          "gasto_key": None},
]


UMBRAL_VENTA_MANUAL_ONLINE = 10.0

# -----------------------------
# Normalizador y Reglas
# -----------------------------
def _norm(s):
    """Normaliza para comparar de forma robusta."""
    return ("" if s is None else str(s)).strip().lower()

def extraer_monto_venta(v):
    """
    Obtiene el monto de la venta desde el modelo con tolerancia a nombres distintos.
    Ajusta el orden si tu campo principal se llama diferente.
    """
    for attr in ("total_venta", "venta_total", "venta_neta_iva", "monto_total", "total"):
        val = getattr(v, attr, None)
        if val is not None:
            try:
                return float(val)
            except (TypeError, ValueError):
                pass
    return 0.0

# -----------------------------
# Reglas de clasificación
# -----------------------------
def es_shopify(n: str) -> bool:
    return "shopify" in n

def es_ubereats(n: str) -> bool:
    n_ns = n.replace(" ", "")
    return n_ns == "ubereats" or ("uber" in n and "eats" in n)

def es_mercadolibre(n: str) -> bool:
    n_ns = n.replace(" ", "")
    return n_ns == "mercadolibre" or ("mercado" in n and "libre" in n)

def clasificar_tienda(comprador, monto=None):
    """
    - 'Venta Manual' => 'ventamanual' (siempre)
    - Shopify        => 'online'
    - Uber Eats      => 'ubereats'
    - Mercado Libre  => 'mercadolibre'
    - Falabella      => 'falabella'
    - Paris          => 'paris'
    - Resto          => 'otros'
    """
    n = _norm(comprador)
    if not n:
        return "otros"

    # Venta Manual SIEMPRE a su propia fila
    if n.replace(" ", "") == "ventamanual":
        return "ventamanual"

    if es_shopify(n):
        return "online"
    if es_ubereats(n):
        return "ubereats"
    if es_mercadolibre(n):
        return "mercadolibre"
    if "falabella" in n:
        return "falabella"
    if "paris" in n:
        return "paris"
    return "otros"

def construir_resumen_tiendas(qs):
    totales = {t["key"]: {"monto": 0.0, "cantidad": 0} for t in TIENDAS}

    for v in qs:
        monto = extraer_monto_venta(v)
        key = clasificar_tienda(getattr(v, "comprador", ""), monto)
        if key not in totales:
            key = "otros"
        totales[key]["monto"] += monto or 0.0
        totales[key]["cantidad"] += 1

    rows = []
    for t in TIENDAS:
        k = t["key"]
        rows.append({
            "key": k,
            "tienda": t["nombre"],
            "ventas": round(totales[k]["monto"], 2),
            "cantidad": totales[k]["cantidad"],
            "gasto_key": t["gasto_key"],
        })
    return rows




# -----------------------------
# Matriz de cuentas (gastos)
# -----------------------------
def obtener_matriz_dict():
    """
    Reconstruye los resúmenes y arma un diccionario de saldos por código para
    poder mapear gastos tienda vía 'gasto_key' (p.ej., 'Pe:3010211').
    """
    # Procesos previos (ajusta a lo indispensable si quieres rendimiento)

    debitos_dict = {d.cuenta_debito: float(d.total_debito) for d in ResumenDebito.objects.all()}
    creditos_dict = {c.cuenta_credito: float(c.total_credito) for c in ResumenCredito.objects.all()}

    matriz_dict = {}
    for fila in balance_rows:
        codigo = str(fila["codigo"])
        debito = debitos_dict.get(int(codigo), 0.0)
        credito = creditos_dict.get(int(codigo), 0.0)
        saldo_deudor = debito - credito if debito > credito else 0.0
        saldo_acreedor = credito - debito if credito > debito else 0.0

        # Activo/Pasivo
        if 1010100 <= int(codigo) <= 2040000:
            matriz_dict[f"A:{codigo}"] = saldo_deudor
            matriz_dict[f"P:{codigo}"] = saldo_acreedor
        # Patrimonio/Gastos
        elif 3010100 <= int(codigo) <= 3010300:
            matriz_dict[f"Pe:{codigo}"] = saldo_deudor
            matriz_dict[f"G:{codigo}"] = saldo_acreedor

    return matriz_dict


# -----------------------------
# Acumulador (pasa el monto)
# -----------------------------
def _acumular_por_tienda(queryset_ventas, keys_tiendas):
    datos_por_tienda = {
        t["key"]: {"nombre": t["nombre"], "ventas_netas": 0.0, "costo_producto": 0.0, "gastos_tienda": 0.0}
        for t in keys_tiendas
    }
    total_ventas = total_costo = 0.0

    for v in queryset_ventas:
        monto = extraer_monto_venta(v)                 # ← obtiene total_venta (o equivalente)
        key = clasificar_tienda(v.comprador, monto)    # ← aplica la nueva regla
        if (key is None) or (key not in datos_por_tienda):
            continue

        vn = float(v.venta_neta_iva or 0.0)
        cv = float(v.costo_venta or 0.0)
        datos_por_tienda[key]["ventas_netas"]  += vn
        datos_por_tienda[key]["costo_producto"] += cv
        total_ventas += vn
        total_costo  += cv

    return datos_por_tienda, total_ventas, total_costo


# -----------------------------
# VIEW: Resumen en HTML
# -----------------------------
@staff_member_required
def resumen_ventas_tiendas_view(request):
    year, start_date, end_date = get_panel_date_range(request)
    regenerate_financial_tables(start_date, end_date)

    matriz_dict = obtener_matriz_dict()
    ventas = VentasConsulta.objects.filter(fecha__gte=start_date, fecha__lte=end_date)

    datos_por_tienda, total_ventas, total_costo = _acumular_por_tienda(ventas, TIENDAS)

    # Construye lista final con gastos y KPIs
    lista_final = []
    total_gastos = 0.0
    total_resultado = 0.0

    for tienda in TIENDAS:
        key = tienda["key"]
        gasto_key = tienda.get("gasto_key")
        gastos = matriz_dict.get(gasto_key, 0.0) if gasto_key else 0.0
        datos_por_tienda[key]["gastos_tienda"] = gastos

        ventas_netas = datos_por_tienda[key]["ventas_netas"]
        costo_producto = datos_por_tienda[key]["costo_producto"]
        gastos_tienda = datos_por_tienda[key]["gastos_tienda"]

        resultado_directo = ventas_netas - costo_producto - gastos_tienda
        porcentaje_ventas = (ventas_netas / total_ventas * 100.0) if total_ventas else 0.0
        porcentaje_resultado = (resultado_directo / ventas_netas * 100.0) if ventas_netas else 0.0
        margen_bruto = round((ventas_netas / costo_producto), 1) if costo_producto else 0.0

        lista_final.append({
            "nombre": tienda["nombre"],
            "ventas_netas": ventas_netas,
            "costo_producto": costo_producto,
            "gastos_tienda": gastos_tienda,
            "resultado_directo": resultado_directo,
            "porcentaje_ventas": porcentaje_ventas,
            "porcentaje_resultado": porcentaje_resultado,
            "margen_bruto": margen_bruto,
        })

        total_gastos += gastos_tienda
        total_resultado += resultado_directo

    # Orden: Online primero, luego alfabético
    lista_final.sort(key=lambda r: (r["nombre"] != "Online", r["nombre"]))

    total_margen_bruto = round((total_ventas / total_costo), 1) if total_costo else 0.0
    total_porcentaje_resultado = (total_resultado / total_ventas * 100.0) if total_ventas else 0.0

    return render(request, "bronz_app/resumen_ventas_tiendas.html", {
        "datos_por_tienda": lista_final,
        "total_ventas": total_ventas,
        "total_costo": total_costo,
        "total_gastos": total_gastos,
        "total_resultado": total_resultado,
        "total_margen_bruto": total_margen_bruto,
        "total_porcentaje_resultado": total_porcentaje_resultado,
        "panel_year": year,
        "panel_start_date": start_date,
        "panel_end_date": end_date,
    })


# -----------------------------
# VIEW: Exportación a Excel
# -----------------------------
@staff_member_required
def exportar_resumen_ventas_tiendas_excel(request):
    year, start_date, end_date = get_panel_date_range(request)
    regenerate_financial_tables(start_date, end_date)

    matriz_dict = obtener_matriz_dict()
    ventas = VentasConsulta.objects.filter(fecha__gte=start_date, fecha__lte=end_date)

    datos_por_tienda, total_ventas, total_costo = _acumular_por_tienda(ventas, TIENDAS)

    rows = []
    total_gastos = 0.0
    total_resultado = 0.0

    for tienda in TIENDAS:
        key = tienda["key"]
        gasto_key = tienda.get("gasto_key")
        gastos = matriz_dict.get(gasto_key, 0.0) if gasto_key else 0.0
        datos_por_tienda[key]["gastos_tienda"] = gastos

        ventas_netas = datos_por_tienda[key]["ventas_netas"]
        costo_producto = datos_por_tienda[key]["costo_producto"]
        resultado_directo = ventas_netas - costo_producto - gastos

        margen_bruto = round((ventas_netas / costo_producto), 1) if costo_producto else 0.0

        rows.append({
            "Tienda": tienda["nombre"],
            "Ventas Netas": ventas_netas,
            "Costo Producto": costo_producto,
            "Gastos Tienda": gastos,
            "Resultado Directo": resultado_directo,
            "Margen Bruto": margen_bruto,
        })

        total_gastos += gastos
        total_resultado += resultado_directo

    margen_bruto_total = round((total_ventas / total_costo), 1) if total_costo else 0.0
    rows.append({
        "Tienda": "TOTAL",
        "Ventas Netas": total_ventas,
        "Costo Producto": total_costo,
        "Gastos Tienda": total_gastos,
        "Resultado Directo": total_resultado,
        "Margen Bruto": margen_bruto_total,
    })

    # Online primero; totales siempre al final
    def orden_fila(row):
        return (row["Tienda"] != "Online", row["Tienda"])
    rows = sorted(rows[:-1], key=orden_fila) + [rows[-1]]

    df = pd.DataFrame(rows)

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename=ResumenVentasTiendas_{timezone.now().date()}.xlsx'
    df.to_excel(response, index=False)
    return response


# ============================================
# Vista del Chatbot de productos BRONZ
# ============================================

def chatbot_view(request):
    """Vista para mostrar la interfaz del chatbot."""
    return render(request, "bronz_app/chatbot.html")


# ============================================
# Vista del Dashboard de Ventas Shopify
# ============================================

def ventas_dashboard_view(request):
    """Vista para mostrar el dashboard de ventas Shopify."""
    return render(request, "bronz_app/ventas_dashboard.html")
