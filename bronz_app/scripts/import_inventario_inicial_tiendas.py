# bronz_app/scripts/import_inventario_inicial_tiendas.py

from datetime import datetime
from pathlib import Path

from django.db import transaction
from django.utils import timezone

from openpyxl import load_workbook

from bronz_app.models import Catalogo
from consult_app.models import BodegaTienda, InventarioInicialTiendas


def _norm(s):
    return "" if s is None else str(s).strip()


def _norm_header(h):
    return _norm(h).lower().replace(" ", "").replace("_", "")


def _parse_fecha(val, date_fmt="%Y-%m-%d"):
    if val is None:
        return timezone.now().date()
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, str):
        txt = val.strip()
        for fmt in (date_fmt, "%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%Y/%m/%d"):
            try:
                return datetime.strptime(txt, fmt).date()
            except Exception:
                pass
    # fallback
    return timezone.now().date()


def run_import(
    file_path: str,
    sheet_name: str | None = None,
    create_missing_tiendas: bool = False,
    strict_sku: bool = False,
    date_format: str = "%Y-%m-%d",
):
    """
    Importa Inventario Inicial desde Excel (.xlsx) a consult_app.InventarioInicialTiendas.
    Upsert por (fecha, sku, tienda).

    Retorna: (creados, actualizados, omitidos)
    """
    xlsx_path = Path(file_path)
    if not xlsx_path.exists():
        raise FileNotFoundError(f"No encuentro el archivo: {xlsx_path}")

    wb = load_workbook(filename=str(xlsx_path), data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active

    headers = [cell.value for cell in ws[1]]
    if not headers or all(h is None for h in headers):
        raise ValueError("No encontré cabecera en la primera fila del Excel.")

    header_map = {_norm_header(h): idx for idx, h in enumerate(headers, start=1)}

    def col(*opciones):
        for opt in opciones:
            key = _norm_header(opt)
            if key in header_map:
                return header_map[key]
        return None

    c_fecha   = col("Fecha")
    c_sku     = col("SKU")
    c_tienda  = col("Tienda", "Bodega", "NombreTienda")
    c_cant    = col("Cantidad", "Qty")
    c_coment  = col("Comentario", "Observacion", "Observación")

    faltan = []
    if c_sku is None:    faltan.append("SKU")
    if c_tienda is None: faltan.append("Tienda")
    if c_cant is None:   faltan.append("Cantidad")
    if faltan:
        raise ValueError(f"Faltan columnas obligatorias: {', '.join(faltan)}")

    # --- Leer filas del Excel ---
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r is None or all(v is None for v in r):
            continue

        fecha = _parse_fecha(r[c_fecha-1], date_format) if c_fecha else timezone.now().date()
        sku_txt = _norm(r[c_sku-1]).upper()
        tienda_txt = _norm(r[c_tienda-1])
        cant_val = r[c_cant-1]
        coment = _norm(r[c_coment-1]) if c_coment else ""

        if not sku_txt or not tienda_txt:
            continue

        try:
            cant = int(cant_val) if cant_val is not None else 0
        except Exception:
            raise ValueError(f"Cantidad inválida para SKU {sku_txt} / Tienda {tienda_txt}: {cant_val}")

        if cant < 0:
            # El modelo es PositiveIntegerField: normaliza a 0
            cant = 0

        rows.append((fecha, sku_txt, tienda_txt, cant, coment))

    if not rows:
        return (0, 0, 0)

    # --- Caches de SKUs y Tiendas ---
    skus = sorted({sku for _, sku, _, _, _ in rows})
    tiendas = sorted({t for _, _, t, _, _ in rows})

    sku_qs = Catalogo.objects.filter(sku__in=skus).only("sku", "id")
    sku_cache = {c.sku.upper(): c for c in sku_qs}

    tienda_qs = BodegaTienda.objects.filter(nombre__in=tiendas).only("nombre", "id")
    tienda_cache = {t.nombre: t for t in tienda_qs}

    if create_missing_tiendas:
        faltantes = [t for t in tiendas if t not in tienda_cache]
        if faltantes:
            BodegaTienda.objects.bulk_create(
                [BodegaTienda(nombre=t) for t in faltantes],
                ignore_conflicts=True
            )
            tienda_qs = BodegaTienda.objects.filter(nombre__in=tiendas).only("nombre", "id")
            tienda_cache = {t.nombre: t for t in tienda_qs}

    # Validación previa básica
    if strict_sku:
        missing_sku = [s for s in skus if s not in sku_cache]
        if missing_sku:
            raise ValueError(f"Hay SKU inexistentes en Catálogo: {', '.join(missing_sku[:20])}"
                             + (f" ...(+{len(missing_sku)-20})" if len(missing_sku) > 20 else ""))

    creados = actualizados = omitidos = 0

    # --- Upsert por (fecha, sku, tienda) ---
    with transaction.atomic():
        for fecha, sku_txt, tienda_txt, cant, coment in rows:
            sku_obj = sku_cache.get(sku_txt)
            tienda_obj = tienda_cache.get(tienda_txt)

            if not sku_obj or not tienda_obj:
                omitidos += 1
                continue

            try:
                obj, created = InventarioInicialTiendas.objects.update_or_create(
                    fecha=fecha,
                    sku=sku_obj,
                    tienda=tienda_obj,
                    defaults={"cantidad": cant, "comentario": coment or None},
                )
                if created:
                    creados += 1
                else:
                    actualizados += 1
            except InventarioInicialTiendas.MultipleObjectsReturned:
                qs = InventarioInicialTiendas.objects.filter(
                    fecha=fecha, sku=sku_obj, tienda=tienda_obj
                ).order_by("id")
                keeper = qs.first()
                extras = qs.exclude(id=keeper.id)
                total = cant + sum((e.cantidad or 0) for e in extras)
                extras.delete()
                keeper.cantidad = total
                if coment:
                    keeper.comentario = coment
                keeper.save(update_fields=["cantidad", "comentario"])
                actualizados += 1

    return (creados, actualizados, omitidos)
