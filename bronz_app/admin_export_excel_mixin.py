# bronz_app/admin_export_excel_mixin.py
from django.urls import path
from django.http import HttpResponse
from django.utils.text import capfirst
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
import decimal
class ExportExcelMixin:
    export_excel_filename = None  # Puedes sobreescribirlo en cada admin si quieres
    export_excel_fields = None    # Si None, usa todos los campos del modelo

    def get_urls(self):
        urls = super().get_urls()
        model_name = self.model._meta.model_name
        custom_urls = [
            path(
                f'exportar_excel_{model_name}/',
                self.admin_site.admin_view(self.exportar_excel_view),
                name=f'exportar_excel_{model_name}'
            ),
        ]
        return custom_urls + urls

    def exportar_excel_view(self, request):
        queryset = self.get_queryset(request)
        fields = self.export_excel_fields or [f.name for f in self.model._meta.fields]

        wb = Workbook()
        ws = wb.active
        ws.title = str(capfirst(self.model._meta.verbose_name_plural))[:31]


        # Cabeceras en negrita
        header_font = Font(bold=True)
        ws.append([capfirst(self.model._meta.get_field(f).verbose_name) for f in fields])
        for cell in ws[1]:
            cell.font = header_font

        # Filas con formato numérico

        for row_idx, obj in enumerate(queryset, start=2):
            for col_idx, f in enumerate(fields, start=1):
                value = getattr(obj, f)
                # Convert Decimal a float para Excel
                if isinstance(value, decimal.Decimal):
                    value = float(value)
                if isinstance(value, float):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.number_format = '#,##0.00'
                elif isinstance(value, int):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.number_format = '#,##0'
                else:
                    ws.cell(row=row_idx, column=col_idx, value=str(value) if value is not None else "")

        # Ancho automático de columnas
        for col_idx, _ in enumerate(fields, 1):
            max_length = max(len(str(cell.value)) for cell in ws[get_column_letter(col_idx)])
            ws.column_dimensions[get_column_letter(col_idx)].width = max_length + 2

        model_name = self.model._meta.model_name
        filename = f"{model_name.capitalize()}.xlsx"
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename={filename}'
        wb.save(response)
        return response

    # Agrega la plantilla personalizada al listado del admin
    change_list_template = "admin/bronz_app/export_excel_change_list.html"
