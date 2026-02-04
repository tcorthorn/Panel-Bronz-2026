from django.shortcuts import render
from django.db.models import Q, Sum, F, FloatField, ExpressionWrapper
from django.core.paginator import Paginator
from .models import ProductoRentable
from bronz_app.utils import (regenerar_ventas_consulta,poblar_movimientos_unificados_credito,poblar_movimientos_unificados_debito,
    regenerar_resumenes_credito_debito,)
from django.db.models import Value, IntegerField, Case, When
from django.db.models.functions import Coalesce, Lower, Trim
import pandas as pd
from datetime import date, datetime
from django.http import HttpResponse
from bronz_app.models import Catalogo, EntradaProductos, Ventas 
from .models import BodegaTienda,EnviosATiendas,InventarioInicialTiendas,AjusteInventarioTienda
from django.db.models import Case, When
from decimal import Decimal
from bronz_app.models import VentasConsulta
from pathlib import Path
from django.contrib import messages
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib.auth.decorators import login_required
from django.shortcuts import redirect
from bronz_app.scripts.import_inventario_inicial_tiendas import run_import as importar_inventario_from_xlsx
import decimal as _dec
from consult_app.validar_plan_cuentas import validar_plan_cuentas
from bronz_app.utils import regenerar_resumenes_credito_debito  
from bronz_app.models import MovimientoUnificadoDebito, MovimientoUnificadoCredito

#_____________________________________
# PRODUCTOS MAS RENTABLES
#_____________________________________

def productos_rentables(request):
    # ---- Recalcula automáticamente cada vez que visitas la página ----
    ProductoRentable.objects.all().delete()

    # Procesos previos (puedes incluir solo los necesarios para el financiero)
    regenerar_ventas_consulta()
    poblar_movimientos_unificados_debito()
    poblar_movimientos_unificados_credito()
    regenerar_resumenes_credito_debito()

    ventas = VentasConsulta.objects.annotate(
        venta_neta_total_fila=ExpressionWrapper(F('cantidad') * F('venta_neta_iva'), output_field=FloatField()),
        costo_total_fila=ExpressionWrapper(F('cantidad') * F('costo_promedio_neto'), output_field=FloatField()),
    )

    resumen = ventas.values('codigo_producto', 'categoria', 'producto').annotate(
        cantidad=Sum('cantidad'),
        venta_total=Sum('venta_neta_total_fila'),
        costo_total=Sum('costo_total_fila')
    )

    for v in resumen:
        venta_total = v['venta_total'] or 0
        costo_total = v['costo_total'] or 0
        utilidad = venta_total - costo_total
        # Margen bruto según tu requerimiento: venta_total / costo_total si costo_total > 0
        margen = (venta_total / costo_total) if costo_total > 0 else 0

        ProductoRentable.objects.create(
            codigo_producto=v['codigo_producto'],
            categoria=v['categoria'],
            producto=v['producto'],
            cantidad=v['cantidad'],
            venta_total=venta_total,
            costo_total=costo_total,
            utilidad_bruta_total=utilidad,
            margen_bruto=margen
        )

    # --- Búsqueda ---
    q = request.GET.get('q', '').strip()
    productos = ProductoRentable.objects.all()
    if q:
        productos = productos.filter(
            Q(codigo_producto__icontains=q) |
            Q(categoria__icontains=q) |
            Q(producto__icontains=q)
        )

    # --- Ordenamiento ---
    sort = request.GET.get('sort', 'venta_total')
    direction = request.GET.get('dir', 'desc')
    allowed_sorts = [
        'codigo_producto', 'categoria', 'producto', 'cantidad',
        'venta_total', 'costo_total', 'utilidad_bruta_total', 'margen_bruto'
    ]
    if sort not in allowed_sorts:
        sort = 'venta_total'
    order_by = sort if direction == 'asc' else f'-{sort}'
    productos = productos.order_by(order_by)

    # --- Paginación ---
    paginator = Paginator(productos, 25)  # 25 productos por página
    page_number = request.GET.get('page')
    productos_page = paginator.get_page(page_number)

    # --- Totales de la página actual ---
    total_cantidad = sum([p.cantidad or 0 for p in productos_page])
    total_venta = sum([p.venta_total or 0 for p in productos_page])
    total_costo = sum([p.costo_total or 0 for p in productos_page])
    total_utilidad = sum([p.utilidad_bruta_total or 0 for p in productos_page])

    # --- Totales globales del filtro ---
    global_total = productos.aggregate(
        global_total_cantidad=Sum('cantidad'),
        global_total_venta=Sum('venta_total'),
        global_total_costo=Sum('costo_total'),
        global_total_utilidad=Sum('utilidad_bruta_total'),
    )

    return render(request, 'consult_app/productos_rentables.html', {
        'productos_page': productos_page,
        'q': q,
        'sort': sort,
        'direction': direction,
        'total_cantidad': total_cantidad,
        'total_venta': total_venta,
        'total_costo': total_costo,
        'total_utilidad': total_utilidad,
        'global_total_cantidad': global_total['global_total_cantidad'] or 0,
        'global_total_venta': global_total['global_total_venta'] or 0,
        'global_total_costo': global_total['global_total_costo'] or 0,
        'global_total_utilidad': global_total['global_total_utilidad'] or 0,
    })

def exportar_productos_excel(request):
    productos = ProductoRentable.objects.all().values()
    df = pd.DataFrame(list(productos))
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=productos_rentables.xlsx'
    df.to_excel(response, index=False)
    return response

#_____________________________________
# NUEVA CONSULTA INVENTARIO
#_____________________________________

# consult_app/views.py

from datetime import date, datetime
import decimal as _dec

from django.shortcuts import render, redirect
from django.http import HttpResponse
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.admin.views.decorators import staff_member_required

from django.db.models import Sum, Value, IntegerField, Case, When, Q
from django.db.models.functions import Coalesce, Lower, Trim

# MODELOS (ajusta las rutas si fuese necesario)
from bronz_app.models import Catalogo, EntradaProductos, Ventas
from .models import (
    BodegaTienda, EnviosATiendas, InventarioInicialTiendas, AjusteInventarioTienda
)

# --------------------------------------------------------------------------------------
# CONFIG
# --------------------------------------------------------------------------------------
OFICINA_ID_DEF = 1
BODEGA_ID_DEF  = 5  # Bodega es una tienda más (vende online)

# --------------------------------------------------------------------------------------
# HELPERS
# --------------------------------------------------------------------------------------
def _parse_fecha_corte(request):
    qs = request.GET.get("fecha_corte")
    if not qs:
        return date.today()
    try:
        return datetime.strptime(qs, "%Y-%m-%d").date()
    except ValueError:
        return date.today()

def _to_decimal_safe(x):
    """Convierte a Decimal de forma segura (0 si None o inválido)."""
    if isinstance(x, _dec.Decimal):
        return x
    try:
        return _dec.Decimal(str(0 if x is None else x))
    except Exception:
        return _dec.Decimal(0)

def _totals_template():
    return {
        "stock_inicial": 0,
        "enviado": 0,
        "recibido": 0,
        "ventas": 0,
        "ajustes": 0,
        "stock_actual": 0,
        "valor_inventario": _dec.Decimal(0),
    }

def _sumas_finales(items):
    """
    Suma columnas requeridas sobre una lista de dicts y SIEMPRE devuelve un dict.
    Tolerante a None y a filas que no traigan 'valor_inventario'.
    """
    from decimal import Decimal as _D  # local para evitar choques; usamos _dec en el resto

    tot = {
        "stock_inicial": 0,
        "enviado": 0,
        "recibido": 0,
        "ventas": 0,
        "ajustes": 0,
        "stock_actual": 0,
        "valor_inventario": _D(0),
    }

    if not items:
        return tot

    for it in items:
        if not isinstance(it, dict):
            continue
        tot["stock_inicial"] += int(it.get("stock_inicial") or 0)
        tot["enviado"]       += int(it.get("enviado") or 0)
        tot["recibido"]      += int(it.get("recibido") or 0)
        tot["ventas"]        += int(it.get("ventas") or 0)
        tot["ajustes"]       += int(it.get("ajustes") or 0)
        tot["stock_actual"]  += int(it.get("stock_actual") or 0)

        vi = it.get("valor_inventario", None)
        if vi is None:
            cu = _to_decimal_safe(it.get("costo_unitario"))
            vi = _to_decimal_safe(it.get("stock_actual")) * cu
        else:
            try:
                vi = _D(str(vi))
            except Exception:
                vi = _D(0)
        tot["valor_inventario"] += vi

    return tot


# ---- AJUSTE PACK: BB0003 -> BB0001 + BB0002 ---------------------------------
def _ajustar_pack_bb0003(filas, oficina_id, pack_sku="BB0003", sku_a="BB0001", sku_b="BB0002"):
    """
    Mueve las ventas del SKU pack (BB0003) hacia BB0001 y BB0002 por tienda
    y deja las ventas de BB0003 en 0. Ajusta stock_actual y valor_inventario.
    MUTACIÓN IN-PLACE sobre 'filas'.
    """
    idx = {(row["tienda_id"], row["sku"]): row for row in filas}

    for (tid, sku), r3 in list(idx.items()):
        if sku != pack_sku:
            continue
        v3 = int(r3.get("ventas") or 0)
        if v3 <= 0:
            continue

        # Asegura filas destino por tienda (si no existen, las crea con ceros)
        for base_sku in (sku_a, sku_b):
            if (tid, base_sku) not in idx:
                row_nueva = {
                    "tienda_id": tid,
                    "tienda": r3.get("tienda"),
                    "sku": base_sku,
                    "producto": "",
                    "categoria": r3.get("categoria", ""),
                    "stock_inicial": 0,
                    "enviado": 0,
                    "recibido": 0,
                    "ventas": 0,
                    "ajustes": 0,
                    "stock_actual": 0,
                    "costo_unitario": _to_decimal_safe(0),
                    "valor_inventario": _to_decimal_safe(0),
                }
                filas.append(row_nueva)
                idx[(tid, base_sku)] = row_nueva

        r1 = idx[(tid, sku_a)]
        r2 = idx[(tid, sku_b)]

        # Sumar ventas del pack a ambos SKUs base
        r1["ventas"] = int(r1.get("ventas") or 0) + v3
        r2["ventas"] = int(r2.get("ventas") or 0) + v3

        # En tiendas (no Oficina) las ventas descuentan stock; al pack se le revierte
        if tid != oficina_id:
            r1["stock_actual"] = int(r1.get("stock_actual") or 0) - v3
            r2["stock_actual"] = int(r2.get("stock_actual") or 0) - v3
            r3["stock_actual"] = int(r3.get("stock_actual") or 0) + v3

        # Dejar el pack sin ventas
        r3["ventas"] = 0

        # Recalcular valores inventario
        r1["valor_inventario"] = _to_decimal_safe(r1["stock_actual"]) * _to_decimal_safe(r1["costo_unitario"])
        r2["valor_inventario"] = _to_decimal_safe(r2["stock_actual"]) * _to_decimal_safe(r2["costo_unitario"])
        r3["valor_inventario"] = _to_decimal_safe(r3["stock_actual"]) * _to_decimal_safe(r3["costo_unitario"])



# --------------------------------------------------------------------------------------
# CÁLCULO DE FILAS (DETALLE POR TIENDA/SKU)
# --------------------------------------------------------------------------------------
def _calcular_filas(fecha_corte):
    # Tiendas y nombre por id
    tiendas = list(BodegaTienda.objects.all().values("id", "nombre"))
    tienda_by_id = {t["id"]: t["nombre"] for t in tiendas}

    # Inventario inicial por tienda/sku
    ini_qs = (
        InventarioInicialTiendas.objects
        .filter(fecha__lte=fecha_corte)
        .values("tienda_id", "sku__sku")
        .annotate(stock_inicial=Coalesce(Sum("cantidad"), Value(0, output_field=IntegerField())))
    )
    inicial = {(r["tienda_id"], r["sku__sku"]): r["stock_inicial"] for r in ini_qs}

    # Envíos desde Oficina a tiendas => recibido por tienda
    envios_qs = (
        EnviosATiendas.objects
        .filter(fecha__lte=fecha_corte)
        .values("tienda_bodega_id", "sku__sku")
        .annotate(recibido=Coalesce(Sum("cantidad"), Value(0, output_field=IntegerField())))
    )
    recibido_por_tienda = {(r["tienda_bodega_id"], r["sku__sku"]): r["recibido"] for r in envios_qs}

    # Total Enviado por Oficina (para fila de Oficina)
    enviados_oficina_qs = (
        EnviosATiendas.objects
        .filter(fecha__lte=fecha_corte)
        .values("sku__sku")
        .annotate(enviado=Coalesce(Sum("cantidad"), Value(0, output_field=IntegerField())))
    )
    enviado_oficina_por_sku = {r["sku__sku"]: r["enviado"] for r in enviados_oficina_qs}

    # Entradas en Oficina (recibido_oficina)
    entradas_oficina_qs = (
        EntradaProductos.objects
        .filter(fecha__lte=fecha_corte)
        .values("sku__sku")
        .annotate(entradas=Coalesce(Sum("cantidad_ingresada"), Value(0, output_field=IntegerField())))
    )
    recibido_oficina_por_sku = {r["sku__sku"]: r["entradas"] for r in entradas_oficina_qs}

    # Resolver IDs por nombre (con fallback)
    nombres_ref = ["Oficina", "Falabella", "Tienda3", "Tienda4", "Bodega", "Otro"]
    tienda_id_map = {
        n: BodegaTienda.objects.filter(nombre__iexact=n).values_list("id", flat=True).first()
        for n in nombres_ref
    }
    OFICINA_ID   = tienda_id_map.get("Oficina")   or OFICINA_ID_DEF
    FALABELLA_ID = tienda_id_map.get("Falabella") or 2
    TIENDA3_ID   = tienda_id_map.get("Tienda3")   or 3
    TIENDA4_ID   = tienda_id_map.get("Tienda4")   or 4
    BODEGA_ID    = tienda_id_map.get("Bodega")    or BODEGA_ID_DEF
    OTRO_ID      = tienda_id_map.get("Otro")      or 6

    # Ventas por tienda (clasificación por 'comprador')
    ventas_qs = (
        Ventas.objects
        .filter(fecha__lte=fecha_corte)
        .annotate(comp_norm=Lower(Trim("comprador")))
        .annotate(
            tienda_id=Case(
                When(comprador__isnull=True, then=Value(BODEGA_ID)),
                When(comp_norm__in=["", "nan", "shopify", "uber eats", "ubereats", "mercado libre"], then=Value(BODEGA_ID)),
                When(Q(comp_norm__contains="falabella"), then=Value(FALABELLA_ID)),
                When(comp_norm="tienda3", then=Value(TIENDA3_ID)),
                When(comp_norm="tienda4", then=Value(TIENDA4_ID)),
                default=Value(OTRO_ID),
                output_field=IntegerField(),
            )
        )
        .values("tienda_id", "sku__sku")
        .annotate(ventas=Coalesce(Sum("cantidad"), Value(0, output_field=IntegerField())))
    )
    ventas_por_tienda = {(r["tienda_id"], r["sku__sku"]): r["ventas"] for r in ventas_qs}

    # Ajustes por tienda
    ajustes_qs = (
        AjusteInventarioTienda.objects
        .filter(fecha__lte=fecha_corte)
        .values("tienda_id", "sku__sku")
        .annotate(ajustes=Coalesce(Sum("cantidad"), Value(0, output_field=IntegerField())))
    )
    ajustes_por_tienda = {(r["tienda_id"], r["sku__sku"]): r["ajustes"] for r in ajustes_qs}

    # Universo de claves (tienda, sku)
    claves = set()
    claves.update(inicial.keys())
    claves.update(recibido_por_tienda.keys())
    claves.update(ventas_por_tienda.keys())
    claves.update(ajustes_por_tienda.keys())

    # Asegurar fila de Oficina para cualquier SKU con envíos/entradas
    for sku in set(enviado_oficina_por_sku.keys()).union(recibido_oficina_por_sku.keys()):
        claves.add((OFICINA_ID, sku))

    # Info de producto
    todos_skus = sorted({sku for (_tid, sku) in claves})
    info_prod = {
        c["sku"]: c
        for c in Catalogo.objects.filter(sku__in=todos_skus).values(
            "sku", "producto", "categoria", "costo_promedio_neto"
        )
    }

    # Filas detalle
    filas = []
    for (tid, sku) in sorted(claves, key=lambda x: (x[0], x[1])):
        es_oficina = (tid == OFICINA_ID)
        stock_inicial = int(inicial.get((tid, sku), 0))
        ventas = int(ventas_por_tienda.get((tid, sku), 0))
        ajustes = int(ajustes_por_tienda.get((tid, sku), 0))

        if es_oficina:
            recibido = int(recibido_oficina_por_sku.get(sku, 0))  # entradas
            enviado  = int(enviado_oficina_por_sku.get(sku, 0))   # envíos a tiendas
            stock_actual = stock_inicial + recibido - enviado + ajustes
        else:
            recibido = int(recibido_por_tienda.get((tid, sku), 0))  # envíos desde Oficina
            enviado  = 0
            stock_actual = stock_inicial + recibido - ventas + ajustes

        prod = info_prod.get(sku, {})
        costo_unitario = _to_decimal_safe(prod.get("costo_promedio_neto"))
        valor_inventario = _to_decimal_safe(stock_actual) * costo_unitario

        filas.append({
            "tienda_id": tid,
            "tienda": tienda_by_id.get(tid, f"ID {tid}"),
            "sku": sku,
            "producto": prod.get("producto") or "",
            "categoria": prod.get("categoria") or "",
            "stock_inicial": stock_inicial,
            "enviado": enviado,
            "recibido": recibido,
            "ventas": ventas,
            "ajustes": ajustes,
            "stock_actual": stock_actual,
            "costo_unitario": costo_unitario,
            "valor_inventario": valor_inventario,
        })

    # Devolver filas, tiendas y IDs reales (para la plantilla)
    return filas, tiendas, {"OFICINA_ID": OFICINA_ID, "BODEGA_ID": BODEGA_ID}

# --------------------------------------------------------------------------------------
# VIEW PRINCIPAL
# --------------------------------------------------------------------------------------
def informe_inventario_tiendas(request):
    # 0) Context base siempre definido (evita UnboundLocalError)
    context = {
        "fecha_corte": None,
        "OFICINA_ID": OFICINA_ID_DEF,
        "BODEGA_ID":  BODEGA_ID_DEF,
        "tiendas": [],
        "filas": [],
        "consolidado": [],
        "sku_list": [],
        "categoria_list": [],
        "totales_detalle": {"stock_inicial":0,"enviado":0,"recibido":0,"ventas":0,"ajustes":0,"stock_actual":0,"valor_inventario":_to_decimal_safe(0)},
        "totales_consolidado": {"stock_inicial":0,"enviado":0,"recibido":0,"ventas":0,"ajustes":0,"stock_actual":0,"valor_inventario":_to_decimal_safe(0)},
        "total_detalle_stock": 0,
        "total_detalle_valor": _to_decimal_safe(0),
        "total_consol_stock": 0,
        "total_consol_valor": _to_decimal_safe(0),
        "ocultar_enviado_consol": True,
        "titulo_consolidado": "Consolidado por SKU (Recibido = Entradas de Productos)",
        "etiqueta_recibido_consol": "Entradas de Productos",
    }

    # 1) Datos base
    fecha_corte = _parse_fecha_corte(request)
    context["fecha_corte"] = fecha_corte

    res = _calcular_filas(fecha_corte)  # debe devolver (filas, tiendas, ids)
    try:
        filas, tiendas, ids = res
    except ValueError:
        # compatibilidad si devuelve solo 2
        filas, tiendas = res
        ids = {"OFICINA_ID": OFICINA_ID_DEF, "BODEGA_ID": BODEGA_ID_DEF}

    # 2) Ajuste del pack (antes del consolidado y de los totales)
    _ajustar_pack_bb0003(filas, ids["OFICINA_ID"])

    # 3) Consolidado por SKU (Recibido = entradas de productos)
    entradas_qs = (
        EntradaProductos.objects
        .filter(fecha__lte=fecha_corte)
        .values("sku__sku")
        .annotate(entradas=Coalesce(Sum("cantidad_ingresada"), Value(0, output_field=IntegerField())))
    )
    entradas_por_sku = {r["sku__sku"]: r["entradas"] for r in entradas_qs}

    consolidado_map = {}
    for f in filas:
        sku = f["sku"]
        agg = consolidado_map.get(sku)
        if not agg:
            agg = consolidado_map[sku] = {
                "sku": sku,
                "producto": f.get("producto",""),
                "categoria": f.get("categoria",""),
                "stock_inicial": 0,
                "recibido": 0,    # se setea con entradas_por_sku
                "ventas": 0,
                "ajustes": 0,
                "stock_actual": 0,
                "costo_unitario": f.get("costo_unitario", _to_decimal_safe(0)),
                "valor_inventario": _to_decimal_safe(0),
            }
        agg["stock_inicial"] += int(f.get("stock_inicial") or 0)
        agg["ventas"]        += int(f.get("ventas") or 0)
        agg["ajustes"]       += int(f.get("ajustes") or 0)
        agg["stock_actual"]  += int(f.get("stock_actual") or 0)

    for agg in consolidado_map.values():
        agg["recibido"] = int(entradas_por_sku.get(agg["sku"], 0))
        agg["valor_inventario"] = _to_decimal_safe(agg["stock_actual"]) * _to_decimal_safe(agg["costo_unitario"])

    consolidado = sorted(consolidado_map.values(), key=lambda x: x["sku"])

    # 4) Totales (deterministas)
    totales_detalle     = _sumas_finales(filas)
    totales_consolidado = _sumas_finales(consolidado)

    # 5) Completar context (una sola vez)
    context.update({
        "OFICINA_ID": ids["OFICINA_ID"],
        "BODEGA_ID":  ids["BODEGA_ID"],
        "tiendas": tiendas,
        "filas": filas,
        "consolidado": consolidado,
        "sku_list": sorted({f["sku"] for f in filas}),
        "categoria_list": sorted({f["categoria"] for f in filas if f.get("categoria")}),
        "totales_detalle": totales_detalle,
        "totales_consolidado": totales_consolidado,
        "total_detalle_stock": totales_detalle["stock_actual"],
        "total_detalle_valor": totales_detalle["valor_inventario"],
        "total_consol_stock":  totales_consolidado["stock_actual"],
        "total_consol_valor":  totales_consolidado["valor_inventario"],
    })

    # 6) Render
    return render(request, "consult_app/inventario_tiendas.html", context)


# --------------------------------------------------------------------------------------
# EXPORTACIÓN A EXCEL
# --------------------------------------------------------------------------------------
def exportar_inventario_tiendas_excel(request):
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    fecha_corte = _parse_fecha_corte(request)
    filas, *_ = _calcular_filas(fecha_corte)  # toma solo 'filas'

    # --- AJUSTE FINAL PACK BB0003 ---
    _ajustar_pack_bb0003(filas, ids["OFICINA_ID"])

    # Consolidado por SKU (Entradas de productos)
    entradas_qs = (
        EntradaProductos.objects
        .filter(fecha__lte=fecha_corte)
        .values("sku__sku")
        .annotate(entradas=Coalesce(Sum("cantidad_ingresada"), Value(0, output_field=IntegerField())))
    )
    entradas_por_sku = {r["sku__sku"]: r["entradas"] for r in entradas_qs}

    consolidado_map = {}
    for f in filas:
        sku = f["sku"]
        if sku not in consolidado_map:
            consolidado_map[sku] = {
                "sku": sku,
                "producto": f["producto"],
                "categoria": f["categoria"],
                "stock_inicial": 0,
                "recibido": 0,
                "ventas": 0,
                "ajustes": 0,
                "stock_actual": 0,
                "costo_unitario": f.get("costo_unitario", _dec.Decimal(0)),
                "valor_inventario": _dec.Decimal(0),
            }
        agg = consolidado_map[sku]
        agg["stock_inicial"] += int(f["stock_inicial"])
        agg["ventas"]        += int(f["ventas"])
        agg["ajustes"]       += int(f["ajustes"])
        agg["stock_actual"]  += int(f["stock_actual"])

    for agg in consolidado_map.values():
        agg["recibido"] = int(entradas_por_sku.get(agg["sku"], 0))
        agg["valor_inventario"] = _to_decimal_safe(agg["stock_actual"]) * _to_decimal_safe(agg["costo_unitario"])

    consolidado = sorted(consolidado_map.values(), key=lambda x: x["sku"])

    # Workbook
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Detalle por tienda"
    ws2 = wb.create_sheet("Consolidado por SKU")

    # Hoja 1: Detalle
    headers1 = [
        "Tienda", "SKU", "Producto", "Categoría",
        "Stock Inicial", "Enviado (Oficina)", "Recibido", "Ventas", "Ajustes",
        "Stock Actual", "Costo Unitario", "Valor Inventario"
    ]
    ws1.append(headers1)
    for f in filas:
        ws1.append([
            f["tienda"], f["sku"], f["producto"], f["categoria"],
            f["stock_inicial"], f["enviado"], f["recibido"],
            f["ventas"], f["ajustes"], f["stock_actual"],
            float(_to_decimal_safe(f["costo_unitario"])),
            float(_to_decimal_safe(f["valor_inventario"])),
        ])

    # Totales (Hoja 1): columnas numéricas 5,6,7,8,9,10,12
    last_row_1 = ws1.max_row
    totals_row_1 = last_row_1 + 1
    ws1.cell(row=totals_row_1, column=4, value="TOTALES")
    for col_idx in [5, 6, 7, 8, 9, 10, 12]:
        col_letter = get_column_letter(col_idx)
        ws1.cell(row=totals_row_1, column=col_idx, value=f"=SUM({col_letter}2:{col_letter}{last_row_1})")

    # Hoja 2: Consolidado
    headers2 = [
        "SKU", "Producto", "Categoría",
        "Stock Inicial", "Entradas de Productos", "Ventas", "Ajustes",
        "Stock Actual", "Costo Unitario", "Valor Inventario"
    ]
    ws2.append(headers2)
    for r in consolidado:
        ws2.append([
            r["sku"], r["producto"], r["categoria"],
            r["stock_inicial"], r["recibido"],
            r["ventas"], r["ajustes"], r["stock_actual"],
            float(_to_decimal_safe(r["costo_unitario"])),
            float(_to_decimal_safe(r["valor_inventario"])),
        ])

    # Totales (Hoja 2): columnas numéricas 4,5,6,7,8,10 (no sumamos costo unitario)
    last_row_2 = ws2.max_row
    totals_row_2 = last_row_2 + 1
    ws2.cell(row=totals_row_2, column=3, value="TOTALES")
    for col_idx in [4, 5, 6, 7, 8, 10]:
        col_letter = get_column_letter(col_idx)
        ws2.cell(row=totals_row_2, column=col_idx, value=f"=SUM({col_letter}2:{col_letter}{last_row_2})")

    # Autofit básico
    from openpyxl.utils import get_column_letter as _gcl
    for ws in (ws1, ws2):
        for col in ws.columns:
            max_len = 0
            col_letter = _gcl(col[0].column)
            for cell in col:
                try:
                    max_len = max(max_len, len(str(cell.value)) if cell.value is not None else 0)
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = min(max(10, max_len + 2), 40)

    file_name = f"inventario_tiendas_{fecha_corte.isoformat()}.xlsx"
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename=\"{file_name}\"'
    wb.save(response)
    return response

# --------------------------------------------------------------------------------------
# IMPORTACIÓN INVENTARIO INICIAL (opcional)
# --------------------------------------------------------------------------------------
DEFAULT_XLSX_PATH = r"C:\Users\tcort\OneDrive\BRONZ\Django-Bronz\Otros\Inventario Inicial Tiendas.xlsx"
DEFAULT_SHEET = None  # o "Hoja1"

@login_required
@staff_member_required
def importar_inventario_inicial_tiendas(request):
    if request.method != "POST":
        return redirect("home")
    try:
        from .scripts.import_inventario_inicial_tiendas import importar_inventario_from_xlsx
        creados, actualizados, omitidos = importar_inventario_from_xlsx(
            file_path=DEFAULT_XLSX_PATH,
            sheet_name=DEFAULT_SHEET,
            create_missing_tiendas=False,
            strict_sku=False,
            date_format="%Y-%m-%d",
        )
        messages.success(
            request,
            f"Inventario Inicial importado. Creados: {creados} | Actualizados: {actualizados} | Omitidos: {omitidos}"
        )
    except Exception as e:
        messages.error(request, f"Error al importar Inventario Inicial: {e}")
    return redirect("home")




# ----------------------------------------
# IMPORTACION iNVENTARIO INICIAL TIENDAS
# ---------------------------------------

DEFAULT_XLSX_PATH = r"C:\Users\tcort\OneDrive\BRONZ\Django-Bronz\Otros\Inventario Inicial Tiendas.xlsx"
DEFAULT_SHEET = None  # o "Hoja1"

@login_required
@staff_member_required
def importar_inventario_inicial_tiendas(request):
    if request.method != "POST":
        return redirect("home")  # ajusta el nombre de tu URL del Home

    try:
        creados, actualizados, omitidos = importar_inventario_from_xlsx(
            file_path=DEFAULT_XLSX_PATH,
            sheet_name=DEFAULT_SHEET,
            create_missing_tiendas=False,  # cambia a True si quieres crear tiendas faltantes
            strict_sku=False,              # True para abortar si viene un SKU no existente
            date_format="%Y-%m-%d",
        )
        messages.success(
            request,
            f"Inventario Inicial importado. Creados: {creados} | Actualizados: {actualizados} | Omitidos: {omitidos}"
        )
    except Exception as e:
        messages.error(request, f"Error al importar Inventario Inicial: {e}")

    return redirect("home")

# --------------------------------------------------------------------------------------
# VALIDAR PLAN DE CUENTAS
# --------------------------------------------------------------------------------------

from django.shortcuts import render
from django.db.models import Sum
from datetime import date
import bronz_app.models as bronz_models  # ✅ módulo completo, fuera de cualquier función

def validar_plan_cuentas_view(request):
    # (Opcional) recalcular resúmenes antes de validar.
# Si no quieres recalcular cada vez, comenta la línea siguiente.
    try:
        regenerar_resumenes_credito_debito()
    except Exception as e:
        messages.warning(request, f"No se pudo recalcular resúmenes: {e}")

    validar_plan_cuentas(request)
    # Vuelve al Home (ajusta el nombre de la URL según tu proyecto)
    return redirect('home')  # o 'home'


# --------------------------------------------------------------------------------------
# MOVIMIENTO CUENTA CONTABLE
# --------------------------------------------------------------------------------------
from collections import defaultdict
from itertools import zip_longest
from io import BytesIO
from datetime import datetime, date
from decimal import Decimal
from django.db.models import Sum
from django.http import HttpResponse
from django.shortcuts import render
import bronz_app.models as bronz_models  # ✅ módulo completo, fuera de cualquier función


def movimientos_cuenta_view(request):
    cuenta_raw = (request.GET.get("cuenta") or "").strip()
    hubo_busqueda = bool(cuenta_raw)
    error_msg = ""
    rows = []
    total_debitos = 0
    total_creditos = 0
    sal

    cuenta = None
    if cuenta_raw:
        try:
            cuenta = int(cuenta_raw)
        except ValueError:
            error_msg = "El código de cuenta debe ser numérico."

    if cuenta and not error_msg:
        debitos_qs = (
            bronz_models.MovimientoUnificadoDebito.objects
            .filter(cta_debito=cuenta)
            .values("fecha", "cta_debito", "monto_debito", "texto_coment", "tabla_origen")
            .order_by("fecha", "id")
        )
        creditos_qs = (
            bronz_models.MovimientoUnificadoCredito.objects
            .filter(cta_credito=cuenta)
            .values("fecha", "cta_credito", "monto_credito", "texto_coment", "tabla_origen")
            .order_by("fecha", "id")
        )

        total_debitos = (
            bronz_models.MovimientoUnificadoDebito.objects
            .filter(cta_debito=cuenta)
            .aggregate(total=Sum("monto_debito"))["total"] or 0
        )
        total_creditos = (
            bronz_models.MovimientoUnificadoCredito.objects
            .filter(cta_credito=cuenta)
            .aggregate(total=Sum("monto_credito"))["total"] or 0
        )

        # ---- NUEVO: saldo (débitos - créditos) ----
        try:
            saldo_cuenta = (total_debitos or 0) - (total_creditos or 0)
        except TypeError:
            # por si viniera algo que no sea numérico, fuerza a Decimal/float
            from decimal import Decimal
            saldo_cuenta = Decimal(total_debitos or 0) - Decimal(total_creditos or 0)



        # --- Emparejar por fecha y alinear por índice (evita líneas vacías) ---
        deb_por_fecha = defaultdict(list)
        for d in debitos_qs:
            k = d["fecha"] or date.min
            deb_por_fecha[k].append({
                "fecha_debito": d["fecha"],
                "cta_debito": d["cta_debito"],
                "monto_debito": d["monto_debito"],
                "coment_debito": d["texto_coment"],
                "origen_debito": d["tabla_origen"],
            })

        cred_por_fecha = defaultdict(list)
        for c in creditos_qs:
            k = c["fecha"] or date.min
            cred_por_fecha[k].append({
                "fecha_credito": c["fecha"],
                "cta_credito": c["cta_credito"],
                "monto_credito": c["monto_credito"],
                "coment_credito": c["texto_coment"],
                "origen_credito": c["tabla_origen"],
            })

        rows = []
        for k_fecha in sorted(set(deb_por_fecha.keys()) | set(cred_por_fecha.keys())):
            ld = deb_por_fecha.get(k_fecha, [])
            lc = cred_por_fecha.get(k_fecha, [])
            for d, c in zip_longest(ld, lc, fillvalue=None):
                row = {
                    "fecha_debito": None, "cta_debito": None, "monto_debito": None,
                    "coment_debito": None, "origen_debito": None,
                    "fecha_credito": None, "cta_credito": None, "monto_credito": None,
                    "coment_credito": None, "origen_credito": None,
                }
                if d: row.update(d)
                if c: row.update(c)
                rows.append(row)
        # --- fin emparejamiento por fecha ---

    context = {
        "hubo_busqueda": hubo_busqueda,
        "error_msg": error_msg,
        "cuenta_valor": cuenta_raw,
        "rows": rows,
        "total_debitos": total_debitos,
        "total_creditos": total_creditos,
        "saldo_cuenta": saldo_cuenta,   # <-- NUEVO
    }
    return render(request, "consult_app/movimientos_cuenta.html", context)


def movimientos_cuenta_endpoint(request):
    """
    Si ?export=excel -> genera y devuelve XLSX.
    En otro caso, delega a movimientos_cuenta_view (sin tocarla).
    """
    if (request.GET.get("export") or "").lower() != "excel":
        return movimientos_cuenta_view(request)

    # --------- MISMAS consultas que usa la view ---------
    cuenta_raw = (request.GET.get("cuenta") or "").strip()
    if not cuenta_raw.isdigit():
        return HttpResponse("El parámetro 'cuenta' es requerido y debe ser numérico.", status=400)
    cuenta = int(cuenta_raw)

    debitos_qs = (
        bronz_models.MovimientoUnificadoDebito.objects
        .filter(cta_debito=cuenta)
        .values("fecha", "cta_debito", "monto_debito", "texto_coment", "tabla_origen")
        .order_by("fecha", "id")
    )
    creditos_qs = (
        bronz_models.MovimientoUnificadoCredito.objects
        .filter(cta_credito=cuenta)
        .values("fecha", "cta_credito", "monto_credito", "texto_coment", "tabla_origen")
        .order_by("fecha", "id")
    )

    total_debitos = (
        bronz_models.MovimientoUnificadoDebito.objects
        .filter(cta_debito=cuenta)
        .aggregate(total=Sum("monto_debito"))["total"] or 0
    )
    total_creditos = (
        bronz_models.MovimientoUnificadoCredito.objects
        .filter(cta_credito=cuenta)
        .aggregate(total=Sum("monto_credito"))["total"] or 0
    )

    # --- Emparejar por fecha para el Excel (mismo orden que la tabla) ---
    deb_por_fecha = defaultdict(list)
    for d in debitos_qs:
        k = d["fecha"] or date.min
        deb_por_fecha[k].append({
            "fecha_debito": d["fecha"],
            "cta_debito": d["cta_debito"],
            "monto_debito": d["monto_debito"],
            "coment_debito": d["texto_coment"],
            "origen_debito": d["tabla_origen"],
        })

    cred_por_fecha = defaultdict(list)
    for c in creditos_qs:
        k = c["fecha"] or date.min
        cred_por_fecha[k].append({
            "fecha_credito": c["fecha"],
            "cta_credito": c["cta_credito"],
            "monto_credito": c["monto_credito"],
            "coment_credito": c["texto_coment"],
            "origen_credito": c["tabla_origen"],
        })

    rows = []
    for k_fecha in sorted(set(deb_por_fecha.keys()) | set(cred_por_fecha.keys())):
        ld = deb_por_fecha.get(k_fecha, [])
        lc = cred_por_fecha.get(k_fecha, [])
        for d, c in zip_longest(ld, lc, fillvalue=None):
            row = {
                "fecha_debito": None, "cta_debito": None, "monto_debito": None,
                "coment_debito": None, "origen_debito": None,
                "fecha_credito": None, "cta_credito": None, "monto_credito": None,
                "coment_credito": None, "origen_credito": None,
            }
            if d: row.update(d)
            if c: row.update(c)
            rows.append(row)
    # --- fin emparejamiento por fecha ---

    return export_movimientos_excel(
        cuenta=cuenta,
        rows=rows,
        total_debitos=total_debitos,
        total_creditos=total_creditos,
    )


# ---------- Exportador a Excel (compatible con la estructura de rows) ----------
def export_movimientos_excel(*, cuenta, rows, total_debitos, total_creditos):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    def _cell_value(v):
        if v in (None, "—", ""):
            return None
        if isinstance(v, Decimal):
            return float(v)
        if isinstance(v, datetime):
            return datetime(v.year, v.month, v.day, v.hour, v.minute, v.second)
        if isinstance(v, date):
            return v
        return v

    wb = Workbook()
    ws = wb.active
    ws.title = f"Cuenta {cuenta}"

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)
    t = ws.cell(row=1, column=1, value=f"Movimientos por Cuenta {cuenta}")
    t.font = Font(bold=True, size=14)
    t.alignment = Alignment(horizontal="center")
    ws.cell(row=2, column=1, value=f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M')}")

    headers = [
        "Fecha", "Cuenta Débito", "Monto Débito", "Comentario", "Tabla Origen",
        "Fecha", "Cuenta Crédito", "Monto Crédito", "Comentario", "Tabla Origen",
    ]
    header_fill = PatternFill("solid", fgColor="E4EDF4")
    header_font = Font(bold=True)
    center = Alignment(horizontal="center")
    right = Alignment(horizontal="right")
    thin = Side(style="thin", color="DDDDDD")
    border = Border(top=thin, bottom=thin, left=thin, right=thin)

    start_row = 4
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border

    number_fmt = "#,##0"
    r = start_row + 1
    for row in rows:
        values = [
            row["fecha_debito"], row["cta_debito"], row["monto_debito"],
            row["coment_debito"], row["origen_debito"],
            row["fecha_credito"], row["cta_credito"], row["monto_credito"],
            row["coment_credito"], row["origen_credito"],
        ]
        for c, val in enumerate(values, 1):
            cell = ws.cell(row=r, column=c, value=_cell_value(val))
            cell.border = border
            if c in (3, 8):
                cell.number_format = number_fmt
                cell.alignment = right
            elif c in (1, 6):
                cell.alignment = center
        r += 1

    ws.cell(row=r, column=2, value="Σ Débitos").font = Font(bold=True)
    td = ws.cell(row=r, column=3, value=_cell_value(total_debitos))
    td.number_format = number_fmt; td.font = Font(bold=True); td.alignment = right

    ws.cell(row=r, column=7, value="Σ Créditos").font = Font(bold=True)
    tc = ws.cell(row=r, column=8, value=_cell_value(total_creditos))
    tc.number_format = number_fmt; tc.font = Font(bold=True); tc.alignment = right

    # Auto-ancho
    for col in range(1, 11):
        max_len = 0
        for row_cells in ws.iter_rows(min_row=4, max_row=r, min_col=col, max_col=col):
            txt = "" if row_cells[0].value is None else str(row_cells[0].value)
            max_len = max(max_len, len(txt))
        ws.column_dimensions[chr(64 + col)].width = min(max(10, max_len + 2), 45)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    data = buf.getvalue()

    filename = f"movimientos_cuenta_{cuenta}.xlsx"
    resp = HttpResponse(
        data,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="{filename}"; filename*=UTF-8\'\'{filename}'
    resp["Content-Length"] = str(len(data))
    resp["Cache-Control"] = "no-store"
    resp["X-Content-Type-Options"] = "nosniff"
    return resp
